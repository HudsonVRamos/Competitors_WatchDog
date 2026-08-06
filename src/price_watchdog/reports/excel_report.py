"""Gerador de relatório comparativo em Excel com formatação Traffic Light.

Este módulo implementa a geração de relatórios Excel consolidados
ao final de cada ciclo de monitoramento. Utiliza openpyxl para criar
planilhas com formatação condicional por cores (verde/amarelo/vermelho)
indicando a posição competitiva de cada produto.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from price_watchdog.models.entities import PriceCycle, PriceRecord
from price_watchdog.models.intelligence_entities import (
    CompetitorIntelligenceRecord,
)


@dataclass
class ReportRow:
    """Linha individual do relatório comparativo.

    Desacopla a geração do relatório das relações SQLAlchemy,
    permitindo uso independente do contexto de sessão do banco.

    Attributes:
        competitor_name: Nome do concorrente
        product_name: Nome do produto monitorado
        our_price: Nosso preço de referência
        extracted_price: Preço extraído do concorrente
        price_difference: Diferença absoluta (extracted - our)
        price_difference_pct: Diferença percentual
    """

    competitor_name: str
    product_name: str
    our_price: float
    extracted_price: float
    price_difference: float
    price_difference_pct: float


# Cores do Traffic Light Report
_GREEN_FILL = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)
_YELLOW_FILL = PatternFill(
    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
)
_RED_FILL = PatternFill(
    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
)

# Colunas do relatório de preços
_COLUMNS = [
    "Concorrente",
    "Produto",
    "Nosso Preço",
    "Preço Deles",
    "Diferença (R$)",
    "Diferença (%)",
    "Status",
]

# Colunas da aba "Composição de Pacotes"
_COMPOSITION_COLUMNS = [
    "Concorrente",
    "Nome do Pacote",
    "Preço Default",
    "Preço Promocional",
    "Duração Promo (meses)",
    "Canais Lineares",
    "Telas Simultâneas",
    "Fibra (Sim/Não)",
    "Velocidade Fibra (Mbps)",
    "Internet Móvel (Sim/Não)",
    "Velocidade Móvel (Mbps)",
    "Streaming 1",
    "Streaming 2",
    "Streaming 3",
]


class ExcelReportGenerator:
    """Gerador de relatório comparativo em Excel.

    Gera arquivos Excel com formatação Traffic Light para indicar
    a posição competitiva de cada produto em relação aos concorrentes.

    Cores:
        - Verde: somos mais baratos que o concorrente
        - Amarelo: diferença inferior a 5% (atenção)
        - Vermelho: nosso preço é mais de 5% acima do concorrente
    """

    def generate(
        self,
        records: list[PriceRecord],
        cycle: PriceCycle,
        intelligence_records: list[CompetitorIntelligenceRecord]
        | None = None,
    ) -> bytes:
        """Gera arquivo Excel com formatação Traffic Light.

        Filtra apenas records com extraction_status == "success"
        e gera uma planilha com as colunas obrigatórias e formatação
        condicional por cores.

        Quando intelligence_records contém registros com status
        "success", gera abas adicionais de "Composição de Pacotes"
        e "Comunicação Comercial". Se nenhum registro tiver sucesso,
        as abas são omitidas.

        Args:
            records: Lista de PriceRecords do ciclo (pode incluir
                falhas, que serão filtradas).
            cycle: PriceCycle correspondente ao relatório.
            intelligence_records: Lista opcional de
                CompetitorIntelligenceRecords do ciclo.

        Returns:
            Conteúdo do arquivo Excel em bytes.
        """
        rows = self._build_rows(records)
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativo de Preços"

        self._write_header(ws, cycle)
        self._write_data(ws, rows)
        self._adjust_column_widths(ws)

        # Gerar abas de inteligência se houver dados bem-sucedidos
        if intelligence_records:
            has_success = any(
                r.extraction_status == "success"
                for r in intelligence_records
            )
            if has_success:
                self._generate_composition_tab(
                    wb, intelligence_records
                )
                self._generate_communication_tab(
                    wb, intelligence_records
                )

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def generate_from_rows(
        self, rows: list[ReportRow], cycle: PriceCycle
    ) -> bytes:
        """Gera arquivo Excel a partir de ReportRows pré-construídos.

        Útil quando os dados já foram extraídos das entidades
        e não há sessão SQLAlchemy disponível.

        Args:
            rows: Lista de ReportRow com dados do relatório.
            cycle: PriceCycle correspondente ao relatório.

        Returns:
            Conteúdo do arquivo Excel em bytes.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativo de Preços"

        self._write_header(ws, cycle)
        self._write_data(ws, rows)
        self._adjust_column_widths(ws)

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _build_rows(self, records: list[PriceRecord]) -> list[ReportRow]:
        """Converte PriceRecords em ReportRows filtrando apenas sucessos.

        Args:
            records: Lista de PriceRecords (inclui falhas).

        Returns:
            Lista de ReportRow apenas com registros bem-sucedidos.
        """
        rows: list[ReportRow] = []
        for record in records:
            if record.extraction_status != "success":
                continue

            # Acessar nome do concorrente via relationships
            competitor_name = ""
            if (
                record.product_config
                and record.product_config.competitor
            ):
                competitor_name = (
                    record.product_config.competitor.name
                )

            product_name = ""
            if record.product_config:
                product_name = record.product_config.product_name

            rows.append(
                ReportRow(
                    competitor_name=competitor_name,
                    product_name=product_name,
                    our_price=record.our_price,
                    extracted_price=record.extracted_price,
                    price_difference=record.price_difference,
                    price_difference_pct=record.price_difference_pct,
                )
            )
        return rows

    def _write_header(self, ws, cycle: PriceCycle) -> None:
        """Escreve cabeçalho do relatório com informações do ciclo.

        Args:
            ws: Worksheet ativa do openpyxl.
            cycle: PriceCycle com metadados do ciclo.
        """
        # Título do relatório
        ws["A1"] = "Relatório Comparativo de Preços"
        ws["A1"].font = Font(bold=True, size=14)

        # Informações do ciclo
        started = ""
        if cycle.started_at:
            started = cycle.started_at.strftime("%d/%m/%Y %H:%M")
        ws["A2"] = f"Ciclo: {started}"
        ws["A2"].font = Font(size=10, italic=True)

        # Cabeçalhos das colunas (linha 4)
        header_font = Font(bold=True, size=11)
        for col_idx, col_name in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=4, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def _write_data(self, ws, rows: list[ReportRow]) -> None:
        """Escreve dados do relatório com formatação Traffic Light.

        Args:
            ws: Worksheet ativa do openpyxl.
            rows: Lista de ReportRow com os dados.
        """
        for row_idx, row_data in enumerate(rows, start=5):
            ws.cell(
                row=row_idx, column=1, value=row_data.competitor_name
            )
            ws.cell(
                row=row_idx, column=2, value=row_data.product_name
            )
            ws.cell(
                row=row_idx, column=3, value=row_data.our_price
            ).number_format = '#,##0.00'
            ws.cell(
                row=row_idx, column=4, value=row_data.extracted_price
            ).number_format = '#,##0.00'
            ws.cell(
                row=row_idx, column=5, value=row_data.price_difference
            ).number_format = '#,##0.00'
            ws.cell(
                row=row_idx,
                column=6,
                value=row_data.price_difference_pct,
            ).number_format = '0.00"%"'

            # Determinar status e aplicar traffic light
            status = self._get_status(row_data.price_difference_pct)
            ws.cell(row=row_idx, column=7, value=status)

            self._apply_traffic_light(
                ws, row_idx, row_data.price_difference_pct
            )

    def _apply_traffic_light(
        self, worksheet, row: int, pct_diff: float
    ) -> None:
        """Aplica formatação condicional por cores (Traffic Light).

        Regras:
            - Verde: our_price < extracted_price (pct_diff > 0,
              concorrente é mais caro, somos competitivos)
            - Amarelo: diferença absoluta inferior a 5%
              (-5 < pct_diff <= 0, estamos ligeiramente acima)
            - Vermelho: our_price mais de 5% acima do concorrente
              (pct_diff <= -5)

        Nota: pct_diff = (extracted - our) / our * 100
            - Positivo: concorrente cobra mais (somos mais baratos)
            - Negativo: concorrente cobra menos (somos mais caros)

        Args:
            worksheet: Worksheet do openpyxl.
            row: Número da linha a ser formatada.
            pct_diff: Diferença percentual calculada.
        """
        if pct_diff > 0:
            # Concorrente cobra mais — somos competitivos
            fill = _GREEN_FILL
        elif pct_diff > -5:
            # Diferença absoluta inferior a 5% — atenção
            fill = _YELLOW_FILL
        else:
            # Nosso preço mais de 5% acima — não competitivo
            fill = _RED_FILL

        for col in range(1, len(_COLUMNS) + 1):
            worksheet.cell(row=row, column=col).fill = fill

    def _get_status(self, pct_diff: float) -> str:
        """Retorna o texto de status baseado na diferença percentual.

        Args:
            pct_diff: Diferença percentual (extracted - our) / our * 100.

        Returns:
            "Competitivo", "Atenção" ou "Não Competitivo".
        """
        if pct_diff > 0:
            return "Competitivo"
        elif pct_diff > -5:
            return "Atenção"
        else:
            return "Não Competitivo"

    def _adjust_column_widths(self, ws) -> None:
        """Ajusta largura das colunas baseado no conteúdo.

        Args:
            ws: Worksheet do openpyxl.
        """
        min_widths = [18, 25, 14, 14, 16, 14, 18]
        for col_idx, min_width in enumerate(min_widths, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min_width

    def has_successful_intelligence(
        self,
        intelligence_records: list[CompetitorIntelligenceRecord]
        | None,
    ) -> bool:
        """Verifica se há registros de inteligência bem-sucedidos.

        Utilizado pelo CycleConsolidator para decidir se deve
        incluir indicação de falha no email de consolidação.

        Args:
            intelligence_records: Lista de registros de inteligência
                do ciclo.

        Returns:
            True se ao menos um registro tem status "success".
        """
        if not intelligence_records:
            return False
        return any(
            r.extraction_status == "success"
            for r in intelligence_records
        )

    def _generate_communication_tab(
        self,
        wb: Workbook,
        intelligence_records: list[CompetitorIntelligenceRecord],
    ) -> None:
        """Gera aba 'Comunicação Comercial' no relatório Excel.

        Cria uma worksheet com dados de comunicação comercial
        de todos os concorrentes com extração bem-sucedida.
        Uma linha por concorrente.

        Colunas: Concorrente, Palavras-chave, Descrição Banner,
        Resumo Posicionamento.

        Args:
            wb: Workbook do openpyxl onde a aba será adicionada.
            intelligence_records: Lista de
                CompetitorIntelligenceRecords do ciclo.
        """
        ws = wb.create_sheet(title="Comunicação Comercial")

        # Colunas da aba de comunicação comercial
        columns = [
            "Concorrente",
            "Palavras-chave",
            "Descrição Banner",
            "Resumo Posicionamento",
        ]

        # Escrever cabeçalhos
        header_font = Font(bold=True, size=11)
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Filtrar apenas registros com extração bem-sucedida
        success_records = [
            record
            for record in intelligence_records
            if record.extraction_status == "success"
        ]

        # Escrever dados — uma linha por concorrente
        for row_idx, record in enumerate(success_records, start=2):
            # Nome do concorrente
            competitor_name = ""
            if record.competitor:
                competitor_name = record.competitor.name
            ws.cell(
                row=row_idx, column=1, value=competitor_name
            )

            # Palavras-chave separadas por vírgula
            keywords_text = ""
            if record.commercial_keywords:
                keywords_text = ", ".join(
                    record.commercial_keywords
                )
            ws.cell(
                row=row_idx, column=2, value=keywords_text
            )

            # Descrição do banner
            ws.cell(
                row=row_idx,
                column=3,
                value=record.home_banner_description or "",
            )

            # Resumo do posicionamento
            ws.cell(
                row=row_idx,
                column=4,
                value=record.commercial_positioning_summary or "",
            )

        # Ajustar larguras das colunas
        comm_widths = [20, 40, 50, 40]
        for col_idx, width in enumerate(comm_widths, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    def _generate_composition_tab(
        self,
        wb: Workbook,
        intelligence_records: list[CompetitorIntelligenceRecord],
    ) -> None:
        """Gera aba "Composição de Pacotes" no relatório Excel.

        Cria uma worksheet com uma linha por pacote identificado,
        exibindo todos os atributos de composição. Campos null
        resultam em células vazias.

        Filtra apenas registros com extraction_status == "success".

        Args:
            wb: Workbook do openpyxl onde a aba será adicionada.
            intelligence_records: Lista de
                CompetitorIntelligenceRecords do ciclo (pode incluir
                falhas, que serão filtradas).
        """
        ws = wb.create_sheet(title="Composição de Pacotes")

        # Cabeçalhos
        header_font = Font(bold=True, size=11)
        for col_idx, col_name in enumerate(
            _COMPOSITION_COLUMNS, start=1
        ):
            cell = ws.cell(
                row=1, column=col_idx, value=col_name
            )
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Dados — uma linha por pacote
        current_row = 2
        for record in intelligence_records:
            if record.extraction_status != "success":
                continue

            # Obter nome do concorrente via relationship
            competitor_name = ""
            if record.competitor:
                competitor_name = record.competitor.name

            for package in record.packages:
                ws.cell(
                    row=current_row,
                    column=1,
                    value=competitor_name,
                )
                ws.cell(
                    row=current_row,
                    column=2,
                    value=package.plan_name,
                )
                # Preço Default
                if package.default_price is not None:
                    cell = ws.cell(
                        row=current_row,
                        column=3,
                        value=package.default_price,
                    )
                    cell.number_format = '#,##0.00'
                # Preço Promocional
                if package.promotional_price is not None:
                    cell = ws.cell(
                        row=current_row,
                        column=4,
                        value=package.promotional_price,
                    )
                    cell.number_format = '#,##0.00'
                # Duração Promo (meses)
                if package.promotional_period_months is not None:
                    ws.cell(
                        row=current_row,
                        column=5,
                        value=package.promotional_period_months,
                    )
                # Canais Lineares
                if package.linear_channels is not None:
                    ws.cell(
                        row=current_row,
                        column=6,
                        value=package.linear_channels,
                    )
                # Telas Simultâneas
                if package.simultaneous_screens is not None:
                    ws.cell(
                        row=current_row,
                        column=7,
                        value=package.simultaneous_screens,
                    )
                # Fibra (Sim/Não)
                if package.has_fiber is not None:
                    fiber_val = (
                        "Sim" if package.has_fiber else "Não"
                    )
                    ws.cell(
                        row=current_row,
                        column=8,
                        value=fiber_val,
                    )
                # Velocidade Fibra (Mbps)
                if package.fiber_speed_mbps is not None:
                    ws.cell(
                        row=current_row,
                        column=9,
                        value=package.fiber_speed_mbps,
                    )
                # Internet Móvel (Sim/Não)
                if package.has_mobile_internet is not None:
                    mobile_val = (
                        "Sim"
                        if package.has_mobile_internet
                        else "Não"
                    )
                    ws.cell(
                        row=current_row,
                        column=10,
                        value=mobile_val,
                    )
                # Velocidade Móvel (Mbps)
                if package.mobile_speed_mbps is not None:
                    ws.cell(
                        row=current_row,
                        column=11,
                        value=package.mobile_speed_mbps,
                    )
                # Streaming 1
                if package.bundled_streaming_1:
                    ws.cell(
                        row=current_row,
                        column=12,
                        value=package.bundled_streaming_1,
                    )
                # Streaming 2
                if package.bundled_streaming_2:
                    ws.cell(
                        row=current_row,
                        column=13,
                        value=package.bundled_streaming_2,
                    )
                # Streaming 3
                if package.bundled_streaming_3:
                    ws.cell(
                        row=current_row,
                        column=14,
                        value=package.bundled_streaming_3,
                    )

                current_row += 1

        # Ajustar largura das colunas
        composition_widths = [
            18, 22, 14, 16, 18, 15, 17, 14, 20, 18,
            18, 14, 14, 14,
        ]
        for col_idx, width in enumerate(
            composition_widths, start=1
        ):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
