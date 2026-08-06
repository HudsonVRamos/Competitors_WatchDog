"""Property-based tests para inteligência competitiva.

Valida propriedades de corretude para validação de composição de pacotes,
keywords, banner, normalização de streamings e demais funcionalidades
do AI Intelligence Extractor.

Feature: competitor-intelligence-expansion
"""

import pytest
from hypothesis import given, settings, assume
from hypothesis import strategies as st

from price_watchdog.scraper.intelligence_extractor import (
    AIIntelligenceExtractor,
)


# --- Estratégias (generators) para Property 1 ---


def valid_default_price() -> st.SearchStrategy[float]:
    """Gera preços default válidos entre 0.01 e 99999.99."""
    return st.floats(
        min_value=0.01,
        max_value=99999.99,
        allow_nan=False,
        allow_infinity=False,
    )


def invalid_default_price_low() -> st.SearchStrategy[float]:
    """Gera preços default inválidos abaixo de 0.01."""
    return st.floats(
        min_value=-99999.99,
        max_value=0.009,
        allow_nan=False,
        allow_infinity=False,
    )


def invalid_default_price_high() -> st.SearchStrategy[float]:
    """Gera preços default inválidos acima de 99999.99."""
    return st.floats(
        min_value=100000.0,
        max_value=999999.99,
        allow_nan=False,
        allow_infinity=False,
    )


def valid_promotional_period() -> st.SearchStrategy[int]:
    """Gera períodos promocionais válidos entre 1 e 36."""
    return st.integers(min_value=1, max_value=36)


def invalid_promotional_period_low() -> st.SearchStrategy[int]:
    """Gera períodos promocionais inválidos abaixo de 1."""
    return st.integers(min_value=-100, max_value=0)


def invalid_promotional_period_high() -> st.SearchStrategy[int]:
    """Gera períodos promocionais inválidos acima de 36."""
    return st.integers(min_value=37, max_value=1000)


def valid_non_negative_int() -> st.SearchStrategy[int]:
    """Gera inteiros não-negativos para campos como canais, telas, velocidades."""
    return st.integers(min_value=0, max_value=99999)


def invalid_negative_int() -> st.SearchStrategy[int]:
    """Gera inteiros negativos para campos numéricos."""
    return st.integers(min_value=-99999, max_value=-1)


def valid_composition() -> st.SearchStrategy[dict]:
    """Gera uma composição completamente válida com todos os campos.

    Garante que promotional_price <= default_price quando ambos
    estão presentes.
    """
    return st.fixed_dictionaries({
        "default_price": valid_default_price(),
        "promotional_period_months": valid_promotional_period(),
        "linear_channels": valid_non_negative_int(),
        "simultaneous_screens": valid_non_negative_int(),
        "fiber_speed_mbps": valid_non_negative_int(),
        "mobile_speed_mbps": valid_non_negative_int(),
    }).flatmap(lambda d: st.fixed_dictionaries({
        "default_price": st.just(d["default_price"]),
        "promotional_price": st.floats(
            min_value=0.01,
            max_value=d["default_price"],
            allow_nan=False,
            allow_infinity=False,
        ),
        "promotional_period_months": st.just(d["promotional_period_months"]),
        "linear_channels": st.just(d["linear_channels"]),
        "simultaneous_screens": st.just(d["simultaneous_screens"]),
        "fiber_speed_mbps": st.just(d["fiber_speed_mbps"]),
        "mobile_speed_mbps": st.just(d["mobile_speed_mbps"]),
    }))


def valid_composition_with_nulls() -> st.SearchStrategy[dict]:
    """Gera composições válidas com subconjunto aleatório de campos None.

    Campos None são aceitos sem erro pela validação.
    """
    return st.fixed_dictionaries({
        "default_price": st.one_of(valid_default_price(), st.none()),
        "promotional_price": st.none(),  # será ajustado abaixo
        "promotional_period_months": st.one_of(
            valid_promotional_period(), st.none()
        ),
        "linear_channels": st.one_of(valid_non_negative_int(), st.none()),
        "simultaneous_screens": st.one_of(valid_non_negative_int(), st.none()),
        "fiber_speed_mbps": st.one_of(valid_non_negative_int(), st.none()),
        "mobile_speed_mbps": st.one_of(valid_non_negative_int(), st.none()),
    }).flatmap(_adjust_promotional_price)


def _adjust_promotional_price(comp: dict) -> st.SearchStrategy[dict]:
    """Ajusta promotional_price para ser <= default_price quando presente."""
    default_price = comp.get("default_price")
    if default_price is not None and default_price >= 0.01:
        # promotional_price pode ser None ou valor <= default_price
        promo_strategy = st.one_of(
            st.none(),
            st.floats(
                min_value=0.01,
                max_value=default_price,
                allow_nan=False,
                allow_infinity=False,
            ),
        )
    else:
        # Se default_price é None, promotional_price só pode ser None
        # ou valor válido (sem constraint de <=)
        promo_strategy = st.one_of(
            st.none(),
            valid_default_price(),
        )
    return promo_strategy.map(lambda promo: {**comp, "promotional_price": promo})


# --- Property Tests ---


@pytest.mark.property
class TestCompositionValidationProperties:
    """Property-based tests para validação de composição de pacotes.

    Feature: competitor-intelligence-expansion, Property 1:
    Validação de composição de pacotes aceita dados válidos e rejeita inválidos.

    **Validates: Requirements 1.2, 5.2**
    """

    @settings(max_examples=100)
    @given(comp=valid_composition())
    def test_property_1_aceita_composicao_valida_completa(
        self, comp: dict
    ) -> None:
        """Property 1: Composição com todos os campos válidos deve ser aceita.

        Para qualquer composição onde Default_Price está entre 0.01 e
        99999.99, Promotional_Price <= Default_Price e entre 0.01 e
        99999.99, Promotional_Period entre 1 e 36, e campos numéricos
        são inteiros >= 0, a validação deve retornar (True, "").

        **Validates: Requirements 1.2, 5.2**
        """
        extractor = AIIntelligenceExtractor()

        is_valid, reason = extractor._validate_composition(comp)

        assert is_valid is True, (
            f"Validação deveria aceitar composição válida, mas rejeitou. "
            f"Razão: '{reason}'. Dados: {comp}"
        )
        assert reason == "", (
            f"Razão deveria ser vazia para composição válida, "
            f"mas foi: '{reason}'"
        )

    @settings(max_examples=100)
    @given(comp=valid_composition_with_nulls())
    def test_property_1_aceita_composicao_valida_com_nulls(
        self, comp: dict
    ) -> None:
        """Property 1: Composição com campos None deve ser aceita.

        Campos com valor None/null são aceitos sem marcar erro,
        conforme Requirement 1.3.

        **Validates: Requirements 1.2, 5.2**
        """
        extractor = AIIntelligenceExtractor()

        is_valid, reason = extractor._validate_composition(comp)

        assert is_valid is True, (
            f"Validação deveria aceitar composição com campos None, "
            f"mas rejeitou. Razão: '{reason}'. Dados: {comp}"
        )
        assert reason == "", (
            f"Razão deveria ser vazia, mas foi: '{reason}'"
        )

    @settings(max_examples=100)
    @given(
        invalid_price=invalid_default_price_low(),
    )
    def test_property_1_rejeita_default_price_abaixo_minimo(
        self, invalid_price: float
    ) -> None:
        """Property 1: Default_Price < 0.01 deve ser rejeitado.

        **Validates: Requirements 1.2, 5.2**
        """
        extractor = AIIntelligenceExtractor()
        comp = {"default_price": invalid_price}

        is_valid, reason = extractor._validate_composition(comp)

        assert is_valid is False, (
            f"Validação deveria rejeitar default_price={invalid_price} "
            f"(< 0.01), mas aceitou."
        )
        assert reason != "", (
            "Razão de rejeição não deveria ser vazia"
        )

    @settings(max_examples=100)
    @given(
        invalid_price=invalid_default_price_high(),
    )
    def test_property_1_rejeita_default_price_acima_maximo(
        self, invalid_price: float
    ) -> None:
        """Property 1: Default_Price > 99999.99 deve ser rejeitado.

        **Validates: Requirements 1.2, 5.2**
        """
        extractor = AIIntelligenceExtractor()
        comp = {"default_price": invalid_price}

        is_valid, reason = extractor._validate_composition(comp)

        assert is_valid is False, (
            f"Validação deveria rejeitar default_price={invalid_price} "
            f"(> 99999.99), mas aceitou."
        )
        assert reason != "", (
            "Razão de rejeição não deveria ser vazia"
        )

    @settings(max_examples=100)
    @given(
        default_price=valid_default_price(),
    )
    def test_property_1_rejeita_promotional_price_maior_que_default(
        self, default_price: float
    ) -> None:
        """Property 1: Promotional_Price > Default_Price deve ser rejeitado.

        **Validates: Requirements 1.2, 5.2**
        """
        assume(default_price < 99999.0)  # espaço para promo ser maior
        extractor = AIIntelligenceExtractor()
        promo_price = default_price + 0.01
        comp = {
            "default_price": default_price,
            "promotional_price": promo_price,
        }

        is_valid, reason = extractor._validate_composition(comp)

        assert is_valid is False, (
            f"Validação deveria rejeitar promotional_price={promo_price} "
            f"> default_price={default_price}, mas aceitou."
        )
        assert "promotional_price" in reason.lower() or "maior" in reason.lower(), (
            f"Razão deveria mencionar promotional_price, "
            f"mas foi: '{reason}'"
        )

    @settings(max_examples=100)
    @given(
        invalid_period=invalid_promotional_period_low(),
    )
    def test_property_1_rejeita_promotional_period_abaixo_minimo(
        self, invalid_period: int
    ) -> None:
        """Property 1: Promotional_Period < 1 deve ser rejeitado.

        **Validates: Requirements 1.2, 5.2**
        """
        extractor = AIIntelligenceExtractor()
        comp = {"promotional_period_months": invalid_period}

        is_valid, reason = extractor._validate_composition(comp)

        assert is_valid is False, (
            f"Validação deveria rejeitar promotional_period_months="
            f"{invalid_period} (< 1), mas aceitou."
        )
        assert reason != "", (
            "Razão de rejeição não deveria ser vazia"
        )

    @settings(max_examples=100)
    @given(
        invalid_period=invalid_promotional_period_high(),
    )
    def test_property_1_rejeita_promotional_period_acima_maximo(
        self, invalid_period: int
    ) -> None:
        """Property 1: Promotional_Period > 36 deve ser rejeitado.

        **Validates: Requirements 1.2, 5.2**
        """
        extractor = AIIntelligenceExtractor()
        comp = {"promotional_period_months": invalid_period}

        is_valid, reason = extractor._validate_composition(comp)

        assert is_valid is False, (
            f"Validação deveria rejeitar promotional_period_months="
            f"{invalid_period} (> 36), mas aceitou."
        )
        assert reason != "", (
            "Razão de rejeição não deveria ser vazia"
        )

    @settings(max_examples=100)
    @given(
        negative_value=invalid_negative_int(),
        field=st.sampled_from([
            "linear_channels",
            "simultaneous_screens",
            "fiber_speed_mbps",
            "mobile_speed_mbps",
        ]),
    )
    def test_property_1_rejeita_campos_numericos_negativos(
        self, negative_value: int, field: str
    ) -> None:
        """Property 1: Campos numéricos (canais, telas, velocidades) < 0 são rejeitados.

        **Validates: Requirements 1.2, 5.2**
        """
        extractor = AIIntelligenceExtractor()
        comp = {field: negative_value}

        is_valid, reason = extractor._validate_composition(comp)

        assert is_valid is False, (
            f"Validação deveria rejeitar {field}={negative_value} "
            f"(< 0), mas aceitou."
        )
        assert field in reason, (
            f"Razão deveria mencionar o campo '{field}', "
            f"mas foi: '{reason}'"
        )

    @settings(max_examples=100)
    @given(
        valid_value=valid_non_negative_int(),
        field=st.sampled_from([
            "linear_channels",
            "simultaneous_screens",
            "fiber_speed_mbps",
            "mobile_speed_mbps",
        ]),
    )
    def test_property_1_aceita_campos_numericos_nao_negativos(
        self, valid_value: int, field: str
    ) -> None:
        """Property 1: Campos numéricos >= 0 devem ser aceitos.

        **Validates: Requirements 1.2, 5.2**
        """
        extractor = AIIntelligenceExtractor()
        comp = {field: valid_value}

        is_valid, reason = extractor._validate_composition(comp)

        assert is_valid is True, (
            f"Validação deveria aceitar {field}={valid_value} "
            f"(>= 0), mas rejeitou. Razão: '{reason}'"
        )


# --- Property 5: Truncamento de banner description a 500 caracteres ---


@pytest.mark.property
class TestBannerTruncationProperty:
    """Property 5: Truncamento de banner description a 500 caracteres.

    Feature: competitor-intelligence-expansion, Property 5
    """

    @settings(max_examples=100)
    @given(description=st.text(min_size=0, max_size=2000))
    def test_property_5_resultado_sempre_max_500_chars(
        self, description: str
    ) -> None:
        """Property 5: Resultado sempre tem no máximo 500 caracteres.

        Para qualquer string de descrição de banner retornada pelo
        Bedrock, o resultado persistido SHALL ter no máximo 500
        caracteres.

        **Validates: Requirements 2.4**
        """
        extractor = AIIntelligenceExtractor()

        result = extractor._validate_banner(description)

        assert len(result) <= 500, (
            f"Banner deveria ter no máximo 500 chars, mas tem "
            f"{len(result)}. Input tinha {len(description)} chars."
        )

    @settings(max_examples=100)
    @given(description=st.text(min_size=0, max_size=2000))
    def test_property_5_resultado_e_prefixo_do_input(
        self, description: str
    ) -> None:
        """Property 5: Resultado é prefixo do input (início preservado).

        Para qualquer string de descrição de banner, o resultado
        truncado SHALL ser um prefixo da string original, garantindo
        que o início do conteúdo não é perdido.

        **Validates: Requirements 2.4**
        """
        extractor = AIIntelligenceExtractor()

        result = extractor._validate_banner(description)

        assert description.startswith(result), (
            f"Resultado deveria ser prefixo do input. "
            f"Input: '{description[:60]}...', "
            f"Resultado: '{result[:60]}...'"
        )

    @settings(max_examples=100)
    @given(description=st.text(min_size=0, max_size=500))
    def test_property_5_input_menor_igual_500_nao_trunca(
        self, description: str
    ) -> None:
        """Property 5: Se input <= 500 chars, resultado igual ao input.

        Para qualquer string com no máximo 500 caracteres, o
        resultado SHALL ser idêntico ao input (sem truncamento).

        **Validates: Requirements 2.4**
        """
        extractor = AIIntelligenceExtractor()

        result = extractor._validate_banner(description)

        assert result == description, (
            f"Input com {len(description)} chars (<= 500) deveria "
            f"retornar inalterado, mas resultado difere. "
            f"Input: '{description[:60]}', "
            f"Result: '{result[:60]}'"
        )


# --- Estratégias para Property 10 ---

# Nomes de serviços de streaming conhecidos (base)
_KNOWN_STREAMING_BASES = [
    "netflix",
    "disney+",
    "paramount+",
    "amazon prime video",
    "globoplay",
    "star+",
    "hbo max",
    "apple tv+",
]

# Nomes genéricos de streaming desconhecidos
_UNKNOWN_STREAMING_BASES = [
    "mubi",
    "crunchyroll",
    "discovery+",
    "starzplay",
    "telecine",
]

# Sufixos de tier que devem ser removidos
_TIER_SUFFIXES = ["Basic", "Premium", "Standard"]


def streaming_base_names() -> st.SearchStrategy[str]:
    """Gera nomes base de streaming (conhecidos + desconhecidos)."""
    return st.sampled_from(
        _KNOWN_STREAMING_BASES + _UNKNOWN_STREAMING_BASES
    )


def tier_suffix() -> st.SearchStrategy[str]:
    """Gera um sufixo de tier para adicionar ao nome do streaming."""
    return st.sampled_from(_TIER_SUFFIXES)


def streaming_name_with_optional_suffix() -> st.SearchStrategy[str]:
    """Gera nome de streaming com sufixo de tier opcional.

    Produz strings como "netflix premium", "Disney+ Basic",
    "hbo max Standard", ou apenas "Globoplay" (sem sufixo).
    """
    base = streaming_base_names()
    suffix = tier_suffix()

    # 50% chance de ter sufixo, 50% sem
    with_suffix = st.tuples(base, suffix).map(
        lambda t: f"{t[0]} {t[1]}"
    )
    without_suffix = base

    return st.one_of(with_suffix, without_suffix)


def streaming_lists() -> st.SearchStrategy[list[str]]:
    """Gera listas de 0 a 10 nomes de streaming com sufixos variados."""
    return st.lists(
        streaming_name_with_optional_suffix(),
        min_size=0,
        max_size=10,
    )


# --- Property 10 Tests ---


@pytest.mark.property
class TestStreamingNormalizationProperty:
    """Property 10: Normalização de nomes de streaming.

    Truncamento a 3 e remoção de sufixos de tier.

    Feature: competitor-intelligence-expansion, Property 10
    """

    @settings(max_examples=100)
    @given(streamings=streaming_lists())
    def test_property_10_resultado_max_3_itens(
        self, streamings: list[str]
    ) -> None:
        """Property 10: Resultado sempre tem no máximo 3 itens.

        Para qualquer lista de 0 a 10 nomes de streaming,
        o resultado SHALL ter no máximo 3 itens.

        **Validates: Requirements 9.2**
        """
        extractor = AIIntelligenceExtractor()

        result = extractor._normalize_streamings(streamings)

        assert len(result) <= 3, (
            f"Resultado deveria ter no máximo 3 streamings, "
            f"mas tem {len(result)}. "
            f"Input ({len(streamings)} itens): {streamings}"
        )

    @settings(max_examples=100)
    @given(streamings=streaming_lists())
    def test_property_10_lista_vazia_retorna_vazia(
        self, streamings: list[str]
    ) -> None:
        """Property 10: Lista vazia de input retorna lista vazia.

        Se nenhum streaming for fornecido (lista vazia),
        o resultado SHALL ser lista vazia.

        **Validates: Requirements 9.2**
        """
        extractor = AIIntelligenceExtractor()

        if len(streamings) == 0:
            result = extractor._normalize_streamings(streamings)
            assert result == [], (
                f"Input vazio deveria retornar lista vazia, "
                f"mas retornou: {result}"
            )

    @settings(max_examples=100)
    @given(streamings=streaming_lists())
    def test_property_10_resultado_sem_sufixos_tier(
        self, streamings: list[str]
    ) -> None:
        """Property 10: Nenhum item do resultado contém sufixos de tier.

        Para qualquer lista de streaming com possíveis sufixos
        (Basic, Premium, Standard), cada item do resultado SHALL
        não conter esses sufixos.

        **Validates: Requirements 9.4**
        """
        extractor = AIIntelligenceExtractor()

        result = extractor._normalize_streamings(streamings)

        for item in result:
            item_lower = item.lower()
            for suffix in _TIER_SUFFIXES:
                # Verificar que o sufixo não aparece como
                # palavra separada no final do nome
                suffix_lower = suffix.lower()
                assert not item_lower.endswith(
                    f" {suffix_lower}"
                ), (
                    f"Item '{item}' ainda contém sufixo de tier "
                    f"'{suffix}'. Input: {streamings}"
                )

    @settings(max_examples=100)
    @given(streamings=streaming_lists())
    def test_property_10_servicos_conhecidos_capitalizacao_oficial(
        self, streamings: list[str]
    ) -> None:
        """Property 10: Serviços conhecidos recebem capitalização oficial.

        Para serviços de streaming conhecidos (Netflix, Disney+,
        Paramount+, etc.), o resultado SHALL usar a capitalização
        oficial independente do case do input.

        **Validates: Requirements 9.4**
        """
        extractor = AIIntelligenceExtractor()

        result = extractor._normalize_streamings(streamings)

        # Mapeia nomes oficiais por lowercase
        known_map = AIIntelligenceExtractor.KNOWN_STREAMINGS

        for item in result:
            item_lower = item.lower()
            if item_lower in known_map:
                expected = known_map[item_lower]
                assert item == expected, (
                    f"Serviço conhecido deveria ter capitalização "
                    f"oficial '{expected}', mas ficou '{item}'. "
                    f"Input: {streamings}"
                )


# --- Estratégias (generators) para Property 4 ---


def valid_keyword() -> st.SearchStrategy[str]:
    """Gera uma keyword válida com 1-50 caracteres não-vazios."""
    return st.text(
        alphabet=st.characters(
            whitelist_categories=("L", "N", "P", "S", "Z"),
            blacklist_characters="\x00",
        ),
        min_size=1,
        max_size=50,
    ).filter(lambda s: s.strip() != "")


def valid_keyword_list() -> st.SearchStrategy[list[str]]:
    """Gera lista de keywords válidas com 3-15 elementos.

    Cada keyword tem no máximo 50 caracteres e é não-vazia
    após strip.
    """
    return st.lists(
        valid_keyword(),
        min_size=3,
        max_size=15,
    )


def too_few_keywords_list() -> st.SearchStrategy[list[str]]:
    """Gera lista de keywords com 0-2 elementos (inválida).

    Listas com menos de 3 keywords devem ser rejeitadas.
    """
    return st.lists(
        valid_keyword(),
        min_size=0,
        max_size=2,
    )


def long_keyword() -> st.SearchStrategy[str]:
    """Gera keyword com mais de 50 caracteres."""
    return st.text(
        alphabet=st.characters(
            whitelist_categories=("L", "N", "P", "S"),
            blacklist_characters="\x00",
        ),
        min_size=51,
        max_size=100,
    ).filter(lambda s: s.strip() != "")


# --- Property 4 Tests ---


@pytest.mark.property
class TestKeywordsValidationProperties:
    """Property-based tests para validação de keywords comerciais.

    Feature: competitor-intelligence-expansion
    Property 4: Validação de keywords aceita listas de 3-15
    com max 50 chars
    """

    @settings(max_examples=100)
    @given(keywords=valid_keyword_list())
    def test_property_4_aceita_lista_valida_de_keywords(
        self, keywords: list[str]
    ) -> None:
        """Property 4: Listas de 3-15 keywords (<=50 chars) aceitas.

        Para qualquer lista de keywords com 3 a 15 elementos onde
        cada elemento tem no máximo 50 caracteres e é não-vazio,
        a validação deve retornar status "identified".

        **Validates: Requirements 2.2, 2.3**
        """
        extractor = AIIntelligenceExtractor()

        validated, status = extractor._validate_keywords(keywords)

        assert status == "identified", (
            f"Lista com {len(keywords)} keywords válidas deveria "
            f"ter status 'identified', mas obteve '{status}'. "
            f"Keywords: {keywords[:3]}..."
        )
        assert len(validated) >= 3, (
            f"Lista validada deveria ter >= 3 keywords, "
            f"mas tem {len(validated)}."
        )
        assert len(validated) <= 15, (
            f"Lista validada deveria ter <= 15 keywords, "
            f"mas tem {len(validated)}."
        )
        # Cada keyword retornada tem no máximo 50 chars
        for kw in validated:
            assert len(kw) <= 50, (
                f"Keyword validada '{kw}' excede 50 chars "
                f"(tem {len(kw)})."
            )

    @settings(max_examples=100)
    @given(keywords=too_few_keywords_list())
    def test_property_4_rejeita_lista_com_menos_de_3_keywords(
        self, keywords: list[str]
    ) -> None:
        """Property 4: Listas <3 keywords retornam não identificado.

        Para qualquer lista com menos de 3 keywords, a validação
        deve retornar status "não identificado" e lista vazia.

        **Validates: Requirements 2.2, 2.3**
        """
        extractor = AIIntelligenceExtractor()

        validated, status = extractor._validate_keywords(keywords)

        assert status == "não identificado", (
            f"Lista com {len(keywords)} keywords deveria ter "
            f"status 'não identificado', mas obteve '{status}'. "
            f"Keywords: {keywords}"
        )
        assert validated == [], (
            f"Lista validada deveria ser vazia para keywords "
            f"insuficientes, mas contém {len(validated)} itens."
        )

    @settings(max_examples=100)
    @given(
        long_kws=st.lists(
            long_keyword(), min_size=3, max_size=10
        )
    )
    def test_property_4_trunca_keywords_longas_a_50_chars(
        self, long_kws: list[str]
    ) -> None:
        """Property 4: Keywords >50 chars são truncadas a 50.

        Para qualquer lista de keywords com mais de 50 caracteres
        cada, a validação deve truncar cada keyword a 50 chars
        e aceitar se a lista resultante tiver >= 3 elementos.

        **Validates: Requirements 2.2, 2.3**
        """
        extractor = AIIntelligenceExtractor()

        validated, status = extractor._validate_keywords(long_kws)

        assert status == "identified", (
            f"Lista com {len(long_kws)} keywords longas "
            f"(truncáveis) deveria ter status 'identified', "
            f"mas obteve '{status}'."
        )
        # Todas as keywords retornadas devem ter <= 50 chars
        for kw in validated:
            assert len(kw) <= 50, (
                f"Keyword '{kw[:20]}...' deveria ter sido "
                f"truncada a 50 chars, mas tem {len(kw)}."
            )

    @settings(max_examples=100)
    @given(
        keywords=st.lists(
            valid_keyword(), min_size=16, max_size=25
        )
    )
    def test_property_4_lista_validada_nao_excede_15(
        self, keywords: list[str]
    ) -> None:
        """Property 4: Lista retornada nunca excede 15 keywords.

        Para qualquer lista de keywords com mais de 15 elementos,
        a lista validada retornada nunca deve ter mais de 15.

        **Validates: Requirements 2.2**
        """
        extractor = AIIntelligenceExtractor()

        validated, status = extractor._validate_keywords(keywords)

        assert status == "identified", (
            f"Lista com {len(keywords)} keywords válidas "
            f"deveria ter status 'identified'."
        )
        assert len(validated) <= 15, (
            f"Lista validada não deveria exceder 15 keywords, "
            f"mas tem {len(validated)}."
        )

    @settings(max_examples=100)
    @given(data=st.data())
    def test_property_4_vazia_ou_none_retorna_nao_identificado(
        self, data: st.DataObject
    ) -> None:
        """Property 4: Entrada vazia/None retorna não identificado.

        Para entradas vazias (lista vazia ou None), a validação
        deve retornar status "não identificado" e lista vazia.

        **Validates: Requirements 2.3**
        """
        extractor = AIIntelligenceExtractor()

        # Testar lista vazia
        validated, status = extractor._validate_keywords([])
        assert status == "não identificado", (
            "Lista vazia deveria retornar 'não identificado'."
        )
        assert validated == [], (
            "Lista vazia deveria retornar lista validada vazia."
        )

        # Testar None (conforme implementação aceita)
        validated_none, status_none = (
            extractor._validate_keywords(
                None  # type: ignore[arg-type]
            )
        )
        assert status_none == "não identificado", (
            "None deveria retornar 'não identificado'."
        )
        assert validated_none == [], (
            "None deveria retornar lista validada vazia."
        )


# --- Estratégias (generators) para Property 2 ---

# Campos opcionais que podem ser null conforme Requirement 1.3
_OPTIONAL_FIELDS = [
    "linear_channels",
    "simultaneous_screens",
    "fiber_speed_mbps",
    "mobile_speed_mbps",
    "promotional_price",
    "promotional_period_months",
    "has_fiber",
    "has_mobile_internet",
]


def plan_name_strategy() -> st.SearchStrategy[str]:
    """Gera nomes de plano válidos (não-vazios)."""
    return st.text(
        alphabet=st.characters(
            whitelist_categories=("L", "N", "Z"),
            blacklist_characters="\x00",
        ),
        min_size=1,
        max_size=50,
    ).filter(lambda s: s.strip() != "")


def null_subset_strategy() -> st.SearchStrategy[set[str]]:
    """Gera subconjunto arbitrário de campos opcionais a serem None."""
    return st.frozensets(
        st.sampled_from(_OPTIONAL_FIELDS),
        min_size=1,
        max_size=len(_OPTIONAL_FIELDS),
    ).map(set)


def package_with_null_fields(
    null_fields: set[str],
) -> dict:
    """Cria pacote dict com campos especificados como None.

    Campos fora do subconjunto recebem valor None também
    para isolar que o comportamento é correto para qualquer
    combinação de nulls.
    """
    pkg: dict = {"plan_name": "Plano Teste"}

    # Definir todos os campos opcionais como None
    for field_name in _OPTIONAL_FIELDS:
        pkg[field_name] = None

    # bundled_streamings sempre como lista vazia (válido)
    pkg["bundled_streamings"] = []

    return pkg


def package_with_mixed_null_fields_strategy() -> (
    st.SearchStrategy[dict]
):
    """Gera pacote com subset aleatório de campos opcionais None.

    Campos que NÃO estão no subconjunto null recebem valores
    válidos para garantir que o pacote não é rejeitado por
    validação de valor (apenas os campos null são testados).
    """
    return st.fixed_dictionaries({
        "plan_name": plan_name_strategy(),
        "null_fields": null_subset_strategy(),
        "default_price": st.one_of(
            st.none(),
            st.floats(
                min_value=10.0,
                max_value=5000.0,
                allow_nan=False,
                allow_infinity=False,
            ),
        ),
    }).map(_build_package_with_nulls)


def _build_package_with_nulls(data: dict) -> dict:
    """Constrói um pacote com campos null conforme subconjunto."""
    plan_name = data["plan_name"]
    null_fields = data["null_fields"]
    default_price = data["default_price"]

    pkg: dict = {
        "plan_name": plan_name,
        "bundled_streamings": [],
    }

    # default_price: pode ser null ou valor válido
    if "default_price" not in null_fields:
        # Valor válido para default_price
        if default_price is not None:
            pkg["default_price"] = default_price
        else:
            pkg["default_price"] = None
    else:
        pkg["default_price"] = None

    # promotional_price
    if "promotional_price" in null_fields:
        pkg["promotional_price"] = None
    else:
        # Se default_price está presente, promo pode ser menor
        if (
            pkg.get("default_price") is not None
            and pkg["default_price"] >= 10.0
        ):
            pkg["promotional_price"] = pkg["default_price"] - 1.0
        else:
            pkg["promotional_price"] = None

    # promotional_period_months
    if "promotional_period_months" in null_fields:
        pkg["promotional_period_months"] = None
    else:
        pkg["promotional_period_months"] = 12

    # Campos inteiros não-negativos
    int_fields = [
        "linear_channels",
        "simultaneous_screens",
        "fiber_speed_mbps",
        "mobile_speed_mbps",
    ]
    for field_name in int_fields:
        if field_name in null_fields:
            pkg[field_name] = None
        else:
            pkg[field_name] = 100

    # Campos booleanos
    bool_fields = ["has_fiber", "has_mobile_internet"]
    for field_name in bool_fields:
        if field_name in null_fields:
            pkg[field_name] = None
        else:
            pkg[field_name] = True

    # Anexar null_fields como metadata para verificação
    pkg["_null_fields"] = null_fields

    return pkg


# --- Property 2 Tests ---


@pytest.mark.property
class TestNullFieldsProperty:
    """Property 2: Campos null não marcam extração como falha.

    Para qualquer resposta JSON válida do Bedrock onde um subconjunto
    arbitrário de campos opcionais é null, a extração SHALL ter status
    diferente de "failed" e os campos null SHALL ser preservados como
    null no resultado.

    Feature: competitor-intelligence-expansion, Property 2

    **Validates: Requirements 1.3**
    """

    @settings(max_examples=100)
    @given(pkg_data=package_with_mixed_null_fields_strategy())
    def test_property_2_pacotes_com_null_nao_sao_rejeitados(
        self, pkg_data: dict
    ) -> None:
        """Property 2: Pacotes com campos null NÃO são rejeitados.

        Para qualquer pacote com plan_name válido e subconjunto
        arbitrário de campos opcionais como None, o pacote SHALL
        aparecer no resultado de _parse_packages (não é descartado).

        **Validates: Requirements 1.3**
        """
        extractor = AIIntelligenceExtractor()

        # Extrair metadata e limpar antes de enviar
        null_fields = pkg_data.pop("_null_fields")
        packages_data = [pkg_data]

        result = extractor._parse_packages(packages_data)

        assert len(result) == 1, (
            f"Pacote com campos null {null_fields} deveria ser "
            f"aceito (aparecer no resultado), mas foi rejeitado. "
            f"Dados: {pkg_data}"
        )

    @settings(max_examples=100)
    @given(pkg_data=package_with_mixed_null_fields_strategy())
    def test_property_2_campos_null_preservados_como_none(
        self, pkg_data: dict
    ) -> None:
        """Property 2: Campos null são preservados como None no resultado.

        Para qualquer pacote onde campos opcionais são null,
        os respectivos atributos no PackageCompositionData
        resultado SHALL ser None.

        **Validates: Requirements 1.3**
        """
        extractor = AIIntelligenceExtractor()

        # Extrair metadata e limpar antes de enviar
        null_fields = pkg_data.pop("_null_fields")
        packages_data = [pkg_data]

        result = extractor._parse_packages(packages_data)

        assert len(result) >= 1, (
            f"Pacote deveria ter sido aceito, mas resultado "
            f"está vazio. Null fields: {null_fields}"
        )

        parsed_pkg = result[0]

        # Verificar que cada campo null está preservado como None
        field_mapping = {
            "linear_channels": "linear_channels",
            "simultaneous_screens": "simultaneous_screens",
            "fiber_speed_mbps": "fiber_speed_mbps",
            "mobile_speed_mbps": "mobile_speed_mbps",
            "promotional_price": "promotional_price",
            "promotional_period_months": "promotional_period_months",
            "has_fiber": "has_fiber",
            "has_mobile_internet": "has_mobile_internet",
        }

        for field_name in null_fields:
            attr_name = field_mapping[field_name]
            actual_value = getattr(parsed_pkg, attr_name)
            assert actual_value is None, (
                f"Campo '{attr_name}' deveria ser None "
                f"(estava null no input), mas é "
                f"{actual_value!r}. "
                f"Null fields: {null_fields}"
            )


# --- Estratégias (generators) para Property 3 ---


def valid_package_dict() -> st.SearchStrategy[dict]:
    """Gera um dicionário de pacote válido para parsing.

    Cada pacote tem plan_name obrigatório e default_price
    válido entre 0.01 e 99999.99. Demais campos são opcionais.
    """
    return st.fixed_dictionaries({
        "plan_name": st.text(
            alphabet=st.characters(
                whitelist_categories=("L", "N", "Z"),
                blacklist_characters="\x00",
            ),
            min_size=1,
            max_size=50,
        ).filter(lambda s: s.strip() != ""),
        "default_price": st.floats(
            min_value=0.01,
            max_value=99999.99,
            allow_nan=False,
            allow_infinity=False,
        ),
        "bundled_streamings": st.lists(
            st.sampled_from([
                "Netflix", "Disney+", "Paramount+",
            ]),
            min_size=0,
            max_size=3,
        ),
    })


def package_list_1_to_25() -> st.SearchStrategy[list[dict]]:
    """Gera listas de 1 a 25 pacotes válidos para teste de limite."""
    return st.lists(
        valid_package_dict(),
        min_size=1,
        max_size=25,
    )


# --- Property 3 Tests ---


@pytest.mark.property
class TestParsePackagesLimitProperty:
    """Property-based tests para parsing de múltiplos pacotes.

    Feature: competitor-intelligence-expansion, Property 3:
    Parsing de múltiplos pacotes com limite de 20.

    Para qualquer lista de pacotes retornada pelo Bedrock com
    tamanho entre 1 e 25, o AI_Intelligence_Extractor SHALL
    parsear cada pacote individualmente e o resultado final
    SHALL conter no máximo 20 pacotes, descartando os excedentes.

    **Validates: Requirements 1.4**
    """

    @settings(max_examples=100)
    @given(packages=package_list_1_to_25())
    def test_property_3_resultado_max_20_pacotes(
        self, packages: list[dict]
    ) -> None:
        """Property 3: Resultado sempre contém no máximo 20 pacotes.

        Para qualquer lista de pacotes com tamanho entre 1 e 25,
        o resultado SHALL ter no máximo 20 pacotes.

        **Validates: Requirements 1.4**
        """
        extractor = AIIntelligenceExtractor()

        result = extractor._parse_packages(packages)

        assert len(result) <= 20, (
            f"Resultado deveria ter no máximo 20 pacotes, "
            f"mas tem {len(result)}. "
            f"Input tinha {len(packages)} pacotes."
        )

    @settings(max_examples=100)
    @given(packages=package_list_1_to_25())
    def test_property_3_resultado_min_input_20_quando_validos(
        self, packages: list[dict]
    ) -> None:
        """Property 3: Resultado é min(input_length, 20) para válidos.

        Quando todos os pacotes no input são válidos (plan_name
        presente e default_price dentro do range), o número de
        pacotes no resultado SHALL ser min(len(input), 20).

        **Validates: Requirements 1.4**
        """
        extractor = AIIntelligenceExtractor()

        result = extractor._parse_packages(packages)

        expected_count = min(len(packages), 20)
        assert len(result) == expected_count, (
            f"Resultado deveria ter {expected_count} pacotes "
            f"(min({len(packages)}, 20)), mas tem "
            f"{len(result)}."
        )

    @settings(max_examples=100)
    @given(packages=package_list_1_to_25())
    def test_property_3_plan_names_correspondem_ao_input(
        self, packages: list[dict]
    ) -> None:
        """Property 3: Cada pacote parseado tem plan_name do input.

        Para qualquer lista de pacotes válidos, cada pacote no
        resultado SHALL ter plan_name correspondente ao pacote
        de mesma posição no input (respeitando o limite de 20).

        **Validates: Requirements 1.4**
        """
        extractor = AIIntelligenceExtractor()

        result = extractor._parse_packages(packages)

        # Verificar que cada plan_name no resultado corresponde
        # ao input (considerando strip aplicado pelo _parse_packages)
        limited_input = packages[:20]
        for i, pkg in enumerate(result):
            expected_name = limited_input[i]["plan_name"].strip()
            assert pkg.plan_name == expected_name, (
                f"Pacote {i}: plan_name deveria ser "
                f"'{expected_name}', mas é '{pkg.plan_name}'."
            )


# --- Estratégias (generators) para Property 14 ---

# Status codes HTTP 4xx que são não-retentáveis (exceto 429)
_NON_RETRYABLE_4XX_CODES = [
    400, 401, 403, 404, 405, 408, 409, 413, 415, 422,
]

# Status codes retentáveis
_RETRYABLE_5XX_CODES = [500, 502, 503, 504]


def non_retryable_4xx_status_code() -> st.SearchStrategy[int]:
    """Gera status codes HTTP 4xx não-retentáveis (exceto 429)."""
    return st.sampled_from(_NON_RETRYABLE_4XX_CODES)


def retryable_5xx_status_code() -> st.SearchStrategy[int]:
    """Gera status codes HTTP 5xx retentáveis."""
    return st.sampled_from(_RETRYABLE_5XX_CODES)


def _make_client_error(status_code: int) -> "ClientError":
    """Cria um botocore ClientError com o status code dado."""
    from botocore.exceptions import ClientError as BotoClientError

    error_response = {
        "Error": {
            "Code": "TestError",
            "Message": f"Error with status {status_code}",
        },
        "ResponseMetadata": {
            "HTTPStatusCode": status_code,
        },
    }
    return BotoClientError(error_response, "InvokeModel")


# --- Property 14 Tests ---


@pytest.mark.property
class TestNonRetryableErrorsProperty:
    """Property 14: Erros não-retentáveis causam falha imediata.

    Para qualquer erro classificado como não-retentável (HTTP 4xx
    exceto 429, erro de validação de schema após esgotamento de
    retries de schema, resposta não-parseável), o
    AI_Intelligence_Extractor SHALL registrar falha imediata com
    zero tentativas adicionais de chamada ao Bedrock.

    Feature: competitor-intelligence-expansion, Property 14

    **Validates: Requirements 10.3**
    """

    @settings(max_examples=100)
    @given(status_code=non_retryable_4xx_status_code())
    def test_property_14_http_4xx_exceto_429_e_non_retryable(
        self, status_code: int
    ) -> None:
        """Property 14: HTTP 4xx (exceto 429) retorna non_retryable.

        Para qualquer status code HTTP 4xx que não seja 429,
        _classify_error SHALL retornar "non_retryable",
        indicando falha imediata sem retry.

        **Validates: Requirements 10.3**
        """
        extractor = AIIntelligenceExtractor()
        error = _make_client_error(status_code)

        result = extractor._classify_error(error)

        assert result == "non_retryable", (
            f"HTTP {status_code} deveria ser classificado como "
            f"'non_retryable', mas retornou '{result}'."
        )

    @settings(max_examples=100)
    @given(data=st.data())
    def test_property_14_json_decode_error_e_non_retryable(
        self, data: st.DataObject
    ) -> None:
        """Property 14: json.JSONDecodeError retorna non_retryable.

        Para qualquer json.JSONDecodeError (resposta
        não-parseável), _classify_error SHALL retornar
        "non_retryable".

        **Validates: Requirements 10.3**
        """
        import json

        extractor = AIIntelligenceExtractor()
        # Gerar diversas mensagens de erro e posições
        msg = data.draw(st.text(min_size=1, max_size=50))
        pos = data.draw(st.integers(min_value=0, max_value=100))
        error = json.JSONDecodeError(msg, "doc_string", pos)

        result = extractor._classify_error(error)

        assert result == "non_retryable", (
            f"JSONDecodeError deveria ser classificado como "
            f"'non_retryable', mas retornou '{result}'. "
            f"Msg: '{msg}', pos: {pos}"
        )

    @settings(max_examples=100)
    @given(
        msg=st.text(min_size=1, max_size=100).filter(
            lambda s: s.strip() != ""
        )
    )
    def test_property_14_value_error_e_non_retryable(
        self, msg: str
    ) -> None:
        """Property 14: ValueError retorna non_retryable.

        Para qualquer ValueError (resposta com valor inválido),
        _classify_error SHALL retornar "non_retryable".

        **Validates: Requirements 10.3**
        """
        extractor = AIIntelligenceExtractor()
        error = ValueError(msg)

        result = extractor._classify_error(error)

        assert result == "non_retryable", (
            f"ValueError('{msg}') deveria ser classificado como "
            f"'non_retryable', mas retornou '{result}'."
        )

    @settings(max_examples=100)
    @given(status_code=retryable_5xx_status_code())
    def test_property_14_http_5xx_e_retryable(
        self, status_code: int
    ) -> None:
        """Property 14: HTTP 5xx retorna retryable (contraste).

        Para qualquer status code HTTP 5xx, _classify_error
        SHALL retornar "retryable", confirmando que apenas 4xx
        (exceto 429) são non_retryable.

        **Validates: Requirements 10.3**
        """
        extractor = AIIntelligenceExtractor()
        error = _make_client_error(status_code)

        result = extractor._classify_error(error)

        assert result == "retryable", (
            f"HTTP {status_code} deveria ser classificado como "
            f"'retryable', mas retornou '{result}'."
        )

    @settings(max_examples=100)
    @given(data=st.data())
    def test_property_14_http_429_e_retryable(
        self, data: st.DataObject
    ) -> None:
        """Property 14: HTTP 429 (throttling) retorna retryable.

        O status 429 é exceção entre os 4xx — deve ser retryable
        (throttling), NÃO non_retryable.

        **Validates: Requirements 10.3**
        """
        extractor = AIIntelligenceExtractor()
        error = _make_client_error(429)

        result = extractor._classify_error(error)

        assert result == "retryable", (
            f"HTTP 429 deveria ser classificado como "
            f"'retryable' (throttling), mas retornou '{result}'."
        )

    @settings(max_examples=100)
    @given(
        msg=st.text(min_size=0, max_size=100)
    )
    def test_property_14_schema_validation_error_e_schema_error(
        self, msg: str
    ) -> None:
        """Property 14: SchemaValidationError retorna schema_error.

        Para qualquer SchemaValidationError, _classify_error SHALL
        retornar "schema_error" (categoria separada de retry de
        schema, distinta de non_retryable e retryable).

        **Validates: Requirements 10.3**
        """
        from price_watchdog.scraper.intelligence_extractor import (
            SchemaValidationError,
        )

        extractor = AIIntelligenceExtractor()
        error = SchemaValidationError(msg)

        result = extractor._classify_error(error)

        assert result == "schema_error", (
            f"SchemaValidationError('{msg}') deveria ser "
            f"classificado como 'schema_error', mas retornou "
            f"'{result}'."
        )

    @settings(max_examples=100)
    @given(data=st.data())
    def test_property_14_timeout_error_e_retryable(
        self, data: st.DataObject
    ) -> None:
        """Property 14: asyncio.TimeoutError retorna retryable.

        Erros de timeout são retentáveis (contraste com
        non_retryable), confirmando a classificação correta.

        **Validates: Requirements 10.3**
        """
        import asyncio

        extractor = AIIntelligenceExtractor()
        error = asyncio.TimeoutError()

        result = extractor._classify_error(error)

        assert result == "retryable", (
            f"asyncio.TimeoutError deveria ser classificado como "
            f"'retryable', mas retornou '{result}'."
        )

    @settings(max_examples=100)
    @given(
        msg=st.text(min_size=0, max_size=50)
    )
    def test_property_14_connection_error_e_retryable(
        self, msg: str
    ) -> None:
        """Property 14: ConnectionError retorna retryable.

        Erros de conexão são retentáveis (contraste com
        non_retryable), confirmando a classificação correta.

        **Validates: Requirements 10.3**
        """
        extractor = AIIntelligenceExtractor()
        error = ConnectionError(msg)

        result = extractor._classify_error(error)

        assert result == "retryable", (
            f"ConnectionError('{msg}') deveria ser classificado "
            f"como 'retryable', mas retornou '{result}'."
        )


# --- Estratégias (generators) para Property 6 ---


def competitor_id_strategy() -> st.SearchStrategy[str]:
    """Gera IDs de concorrentes no formato UUID."""
    return st.uuids().map(str)


def cycle_id_strategy() -> st.SearchStrategy[str]:
    """Gera IDs de ciclo no formato UUID."""
    return st.uuids().map(str)


def extraction_status_strategy() -> st.SearchStrategy[str]:
    """Gera status de extração válidos."""
    return st.sampled_from(["success", "failed", "no_packages_found"])


def intelligence_record_strategy() -> st.SearchStrategy:
    """Gera instâncias mock de CompetitorIntelligenceRecord.

    Retorna dicts com os dados para construir o record,
    pois a instância real requer mapeamento SQLAlchemy.
    """
    return st.fixed_dictionaries({
        "competitor_id": competitor_id_strategy(),
        "cycle_id": cycle_id_strategy(),
        "extraction_status": extraction_status_strategy(),
    })


def record_sequence_strategy() -> st.SearchStrategy[list[dict]]:
    """Gera sequências de 1-5 records para o mesmo competitor_id.

    Simula múltiplos ciclos para um mesmo concorrente,
    onde cada record tem um cycle_id diferente.
    """
    return competitor_id_strategy().flatmap(
        lambda comp_id: st.lists(
            st.fixed_dictionaries({
                "competitor_id": st.just(comp_id),
                "cycle_id": cycle_id_strategy(),
                "extraction_status": extraction_status_strategy(),
            }),
            min_size=1,
            max_size=5,
        )
    )


# --- Property 6 Tests ---


@pytest.mark.property
class TestAppendOnlyPersistenceProperty:
    """Property 6: Persistência append-only preserva registros anteriores.

    Para qualquer sequência de Competitor_Intelligence_Records persistidos
    para um mesmo concorrente ao longo de múltiplos ciclos, cada novo
    registro SHALL ser inserido sem alterar ou remover os registros de
    ciclos anteriores — a contagem total de registros do concorrente
    SHALL crescer monotonicamente.

    Feature: competitor-intelligence-expansion, Property 6

    **Validates: Requirements 3.5**
    """

    @settings(max_examples=100)
    @given(records=record_sequence_strategy())
    def test_property_6_save_record_sempre_usa_session_add(
        self, records: list[dict]
    ) -> None:
        """Property 6: save_record sempre chama session.add (append-only).

        Para qualquer sequência de records do mesmo competitor_id,
        cada chamada a save_record SHALL usar session.add (INSERT),
        garantindo comportamento append-only.

        **Validates: Requirements 3.5**
        """
        import asyncio
        from unittest.mock import AsyncMock, MagicMock, patch
        from contextlib import asynccontextmanager

        from price_watchdog.storage.intelligence_store import (
            IntelligenceStore,
        )
        from price_watchdog.models.intelligence_entities import (
            CompetitorIntelligenceRecord,
        )

        store = IntelligenceStore()

        # Mock session que rastreia chamadas
        mock_session = MagicMock()
        mock_session.add = MagicMock()
        mock_session.merge = MagicMock()
        mock_session.delete = MagicMock()

        @asynccontextmanager
        async def mock_get_session():
            yield mock_session

        with patch(
            "price_watchdog.storage.intelligence_store.get_session",
            mock_get_session,
        ):
            for record_data in records:
                mock_session.reset_mock()

                record = MagicMock(spec=CompetitorIntelligenceRecord)
                record.competitor_id = record_data["competitor_id"]
                record.cycle_id = record_data["cycle_id"]
                record.extraction_status = record_data[
                    "extraction_status"
                ]

                asyncio.run(store.save_record(record))

                # Verifica que session.add foi chamado
                assert mock_session.add.called, (
                    f"save_record deveria chamar session.add para "
                    f"competitor_id={record_data['competitor_id']}, "
                    f"cycle_id={record_data['cycle_id']}, mas não chamou."
                )

    @settings(max_examples=100)
    @given(records=record_sequence_strategy())
    def test_property_6_save_record_nunca_usa_merge_ou_delete(
        self, records: list[dict]
    ) -> None:
        """Property 6: save_record NUNCA chama session.merge ou delete.

        Para qualquer sequência de records, save_record SHALL NEVER
        usar session.merge, session.delete ou session.query(...).update()
        — garantindo que registros anteriores não são alterados.

        **Validates: Requirements 3.5**
        """
        import asyncio
        from unittest.mock import AsyncMock, MagicMock, patch
        from contextlib import asynccontextmanager

        from price_watchdog.storage.intelligence_store import (
            IntelligenceStore,
        )
        from price_watchdog.models.intelligence_entities import (
            CompetitorIntelligenceRecord,
        )

        store = IntelligenceStore()

        # Mock session que rastreia chamadas
        mock_session = MagicMock()
        mock_session.add = MagicMock()
        mock_session.merge = MagicMock()
        mock_session.delete = MagicMock()

        # Mock query para detectar .update()
        mock_query = MagicMock()
        mock_session.query = MagicMock(return_value=mock_query)
        mock_query.filter = MagicMock(return_value=mock_query)
        mock_query.update = MagicMock()

        @asynccontextmanager
        async def mock_get_session():
            yield mock_session

        with patch(
            "price_watchdog.storage.intelligence_store.get_session",
            mock_get_session,
        ):
            for record_data in records:
                record = MagicMock(spec=CompetitorIntelligenceRecord)
                record.competitor_id = record_data["competitor_id"]
                record.cycle_id = record_data["cycle_id"]
                record.extraction_status = record_data[
                    "extraction_status"
                ]

                asyncio.run(store.save_record(record))

            # Após todas as inserções, merge e delete NUNCA devem
            # ter sido chamados
            assert not mock_session.merge.called, (
                f"save_record NÃO deveria chamar session.merge, "
                f"mas chamou {mock_session.merge.call_count} vezes. "
                f"Records: {len(records)}"
            )
            assert not mock_session.delete.called, (
                f"save_record NÃO deveria chamar session.delete, "
                f"mas chamou {mock_session.delete.call_count} vezes. "
                f"Records: {len(records)}"
            )
            assert not mock_query.update.called, (
                f"save_record NÃO deveria chamar query.update(), "
                f"mas chamou {mock_query.update.call_count} vezes. "
                f"Records: {len(records)}"
            )

    @settings(max_examples=100)
    @given(records=record_sequence_strategy())
    def test_property_6_cada_record_gera_session_add_separado(
        self, records: list[dict]
    ) -> None:
        """Property 6: Cada record resulta em chamada separada a session.add.

        Para qualquer sequência de N records para o mesmo competitor_id,
        o total de chamadas a session.add SHALL ser exatamente N —
        um INSERT por record, crescimento monotônico garantido.

        **Validates: Requirements 3.5**
        """
        import asyncio
        from unittest.mock import AsyncMock, MagicMock, patch
        from contextlib import asynccontextmanager

        from price_watchdog.storage.intelligence_store import (
            IntelligenceStore,
        )
        from price_watchdog.models.intelligence_entities import (
            CompetitorIntelligenceRecord,
        )

        store = IntelligenceStore()
        add_call_count = 0

        # Mock session que conta chamadas add
        mock_session = MagicMock()

        def track_add(record):
            nonlocal add_call_count
            add_call_count += 1

        mock_session.add = MagicMock(side_effect=track_add)

        @asynccontextmanager
        async def mock_get_session():
            yield mock_session

        with patch(
            "price_watchdog.storage.intelligence_store.get_session",
            mock_get_session,
        ):
            for record_data in records:
                record = MagicMock(spec=CompetitorIntelligenceRecord)
                record.competitor_id = record_data["competitor_id"]
                record.cycle_id = record_data["cycle_id"]
                record.extraction_status = record_data[
                    "extraction_status"
                ]

                asyncio.run(store.save_record(record))

        expected_adds = len(records)
        assert add_call_count == expected_adds, (
            f"Para {expected_adds} records, deveria haver "
            f"{expected_adds} chamadas a session.add, "
            f"mas houve {add_call_count}. "
            f"Crescimento monotônico violado."
        )


# --- Estratégias (generators) para Property 9 ---


def valid_domain() -> st.SearchStrategy[str]:
    """Gera domínios válidos com pelo menos um ponto."""
    # Componentes de domínio: letras e dígitos, 1-20 chars
    label = st.text(
        alphabet=st.characters(
            whitelist_categories=("L", "N"),
        ),
        min_size=1,
        max_size=20,
    ).filter(lambda s: s.isascii() and s.strip() != "")

    tld = st.sampled_from([
        "com", "net", "org", "br", "io", "dev", "app",
        "com.br", "gov.br", "edu",
    ])

    return st.tuples(label, tld).map(
        lambda t: f"{t[0]}.{t[1]}"
    )


def valid_http_url() -> st.SearchStrategy[str]:
    """Gera URLs válidas com esquema http ou https e domínio válido."""
    scheme = st.sampled_from(["http", "https"])
    domain = valid_domain()
    path = st.sampled_from([
        "", "/", "/page", "/path/to/resource",
        "/home-landing", "/produtos",
    ])

    return st.tuples(scheme, domain, path).map(
        lambda t: f"{t[0]}://{t[1]}{t[2]}"
    )


def invalid_scheme_url() -> st.SearchStrategy[str]:
    """Gera URLs com esquemas inválidos (não http/https)."""
    scheme = st.sampled_from([
        "ftp", "file", "ssh", "telnet", "ws", "wss",
        "mailto", "data", "javascript",
    ])
    domain = valid_domain()

    return st.tuples(scheme, domain).map(
        lambda t: f"{t[0]}://{t[1]}/path"
    )


def url_without_scheme() -> st.SearchStrategy[str]:
    """Gera strings que parecem URLs mas sem esquema http/https."""
    domain = valid_domain()

    return st.one_of(
        # Apenas domínio sem esquema
        domain,
        # Domínio com path sem esquema
        domain.map(lambda d: f"{d}/path"),
        # Com "www." mas sem esquema
        domain.map(lambda d: f"www.{d}"),
    )


def url_without_domain() -> st.SearchStrategy[str]:
    """Gera URLs com esquema válido mas sem domínio válido."""
    return st.sampled_from([
        "http://",
        "https://",
        "http:///path",
        "https:///path",
        "http:// ",
        "https:// /path",
    ])


def url_too_long() -> st.SearchStrategy[str]:
    """Gera URLs válidas porém acima de 2048 caracteres."""
    # Cria URL base válida e adiciona path longo
    return valid_domain().map(
        lambda d: f"https://{d}/{'a' * 2040}"
    )


# --- Property 9 Tests ---


@pytest.mark.property
class TestUrlValidationProperty:
    """Property 9: Validação de URL intelligence_home_url.

    Para qualquer string, a validação SHALL aceitar apenas URLs com
    esquema http ou https seguido de domínio válido, e SHALL rejeitar
    strings que não sejam URLs válidas (sem esquema, esquema inválido,
    sem domínio).

    Feature: competitor-intelligence-expansion, Property 9

    **Validates: Requirements 8.5**
    """

    @settings(max_examples=100)
    @given(url=valid_http_url())
    def test_property_9_aceita_urls_http_https_com_dominio(
        self, url: str
    ) -> None:
        """Property 9: URLs http/https com domínio válido → True.

        Para qualquer URL com esquema http ou https e domínio
        contendo pelo menos um ponto, validate_intelligence_url
        SHALL retornar True.

        **Validates: Requirements 8.5**
        """
        from price_watchdog.registry.competitor_manager import (
            CompetitorManager,
        )

        result = CompetitorManager.validate_intelligence_url(url)

        assert result is True, (
            f"URL válida '{url}' deveria ser aceita (True), "
            f"mas retornou False."
        )

    @settings(max_examples=100)
    @given(url=invalid_scheme_url())
    def test_property_9_rejeita_esquema_invalido(
        self, url: str
    ) -> None:
        """Property 9: URLs com esquema inválido (ftp, file, etc) → False.

        Para qualquer URL com esquema diferente de http/https,
        validate_intelligence_url SHALL retornar False.

        **Validates: Requirements 8.5**
        """
        from price_watchdog.registry.competitor_manager import (
            CompetitorManager,
        )

        result = CompetitorManager.validate_intelligence_url(url)

        assert result is False, (
            f"URL com esquema inválido '{url}' deveria ser "
            f"rejeitada (False), mas retornou True."
        )

    @settings(max_examples=100)
    @given(url=url_without_scheme())
    def test_property_9_rejeita_url_sem_esquema(
        self, url: str
    ) -> None:
        """Property 9: Strings sem esquema http/https → False.

        Para qualquer string que não possua esquema http ou https,
        validate_intelligence_url SHALL retornar False.

        **Validates: Requirements 8.5**
        """
        from price_watchdog.registry.competitor_manager import (
            CompetitorManager,
        )

        result = CompetitorManager.validate_intelligence_url(url)

        assert result is False, (
            f"String sem esquema '{url}' deveria ser "
            f"rejeitada (False), mas retornou True."
        )

    @settings(max_examples=100)
    @given(url=url_without_domain())
    def test_property_9_rejeita_url_sem_dominio(
        self, url: str
    ) -> None:
        """Property 9: URLs com esquema válido mas sem domínio → False.

        Para qualquer URL com esquema http/https mas sem domínio
        válido, validate_intelligence_url SHALL retornar False.

        **Validates: Requirements 8.5**
        """
        from price_watchdog.registry.competitor_manager import (
            CompetitorManager,
        )

        result = CompetitorManager.validate_intelligence_url(url)

        assert result is False, (
            f"URL sem domínio '{url}' deveria ser "
            f"rejeitada (False), mas retornou True."
        )

    @settings(max_examples=100)
    @given(url=url_too_long())
    def test_property_9_rejeita_url_acima_2048_chars(
        self, url: str
    ) -> None:
        """Property 9: URLs com mais de 2048 caracteres → False.

        Para qualquer URL com comprimento > 2048 caracteres,
        validate_intelligence_url SHALL retornar False,
        independente de formato válido.

        **Validates: Requirements 8.5**
        """
        from price_watchdog.registry.competitor_manager import (
            CompetitorManager,
        )

        # Garantir que a URL gerada realmente excede 2048
        assert len(url) > 2048, (
            f"URL gerada deveria exceder 2048 chars, mas tem "
            f"{len(url)}."
        )

        result = CompetitorManager.validate_intelligence_url(url)

        assert result is False, (
            f"URL com {len(url)} chars (> 2048) deveria ser "
            f"rejeitada (False), mas retornou True."
        )

    @settings(max_examples=100)
    @given(data=st.data())
    def test_property_9_rejeita_none_e_vazio(
        self, data: st.DataObject
    ) -> None:
        """Property 9: None e string vazia → False.

        Para entrada None ou string vazia, validate_intelligence_url
        SHALL retornar False.

        **Validates: Requirements 8.5**
        """
        from price_watchdog.registry.competitor_manager import (
            CompetitorManager,
        )

        # Testar None
        result_none = CompetitorManager.validate_intelligence_url(None)
        assert result_none is False, (
            "None deveria retornar False."
        )

        # Testar string vazia
        result_empty = CompetitorManager.validate_intelligence_url("")
        assert result_empty is False, (
            "String vazia deveria retornar False."
        )

        # Testar apenas espaços
        spaces = data.draw(
            st.text(
                alphabet=" \t\n\r",
                min_size=1,
                max_size=10,
            )
        )
        result_spaces = CompetitorManager.validate_intelligence_url(
            spaces
        )
        assert result_spaces is False, (
            f"String de espaços '{spaces!r}' deveria retornar False."
        )


# --- Estratégias (generators) para Property 8 ---

import asyncio
from unittest.mock import AsyncMock, MagicMock, patch

# Tipos de exceção que podem ocorrer na extração de inteligência
_INTELLIGENCE_ERROR_TYPES = [
    TimeoutError,
    RuntimeError,
    ConnectionError,
    ValueError,
    OSError,
    asyncio.TimeoutError,
]


def intelligence_error_type_strategy() -> st.SearchStrategy:
    """Gera tipos de exceção que podem ocorrer na inteligência."""
    return st.sampled_from(_INTELLIGENCE_ERROR_TYPES)


def error_message_strategy() -> st.SearchStrategy[str]:
    """Gera mensagens de erro variadas."""
    return st.text(
        alphabet=st.characters(
            whitelist_categories=("L", "N", "Z", "P"),
            blacklist_characters="\x00",
        ),
        min_size=1,
        max_size=100,
    ).filter(lambda s: s.strip() != "")


def screenshot_bytes_strategy() -> st.SearchStrategy[bytes]:
    """Gera bytes de screenshot simulados (não-vazio)."""
    return st.binary(min_size=10, max_size=200)


# --- Property 8 Tests ---


@pytest.mark.property
class TestFaultIsolationProperty:
    """Property 8: Isolamento — falhas de inteligência não impactam preços.

    Para qualquer falha na extração de inteligência competitiva
    (timeout, resposta inválida, erro de rede), os PriceRecords
    já coletados naquele ciclo para o mesmo concorrente SHALL
    permanecer intactos e inalterados.

    Feature: competitor-intelligence-expansion, Property 8

    **Validates: Requirements 4.3, 10.1**
    """

    def _build_worker_with_failing_extractor(
        self, error: Exception
    ) -> "Worker":
        """Constrói Worker com intelligence_extractor que levanta erro.

        O extractor é mockado para levantar a exceção fornecida,
        simulando falha na extração de inteligência. Os demais
        componentes são mockados com comportamento padrão.
        """
        from price_watchdog.worker.worker import Worker

        consumer = MagicMock()
        scraper = MagicMock()
        comparator = MagicMock()
        price_store = MagicMock()
        screenshot_store = MagicMock()
        alert_service = MagicMock()

        # Intelligence extractor que falha
        intelligence_extractor = MagicMock()
        intelligence_extractor.extract = AsyncMock(
            side_effect=error
        )

        # Intelligence store mockado
        intelligence_store = MagicMock()
        intelligence_store.save_record = AsyncMock()

        worker = Worker(
            consumer=consumer,
            scraper=scraper,
            comparator=comparator,
            price_store=price_store,
            screenshot_store=screenshot_store,
            alert_service=alert_service,
            intelligence_extractor=intelligence_extractor,
            intelligence_store=intelligence_store,
        )

        return worker

    @settings(max_examples=100)
    @given(
        error_type=intelligence_error_type_strategy(),
        error_msg=error_message_strategy(),
        screenshot=screenshot_bytes_strategy(),
        competitor_id=st.uuids().map(str),
        cycle_id=st.uuids().map(str),
    )
    def test_property_8_process_intelligence_nunca_propaga_excecao(
        self,
        error_type: type,
        error_msg: str,
        screenshot: bytes,
        competitor_id: str,
        cycle_id: str,
    ) -> None:
        """Property 8: _process_intelligence NUNCA propaga exceções.

        Para qualquer tipo de falha na inteligência (TimeoutError,
        RuntimeError, ConnectionError, ValueError), o método
        _process_intelligence SHALL completar sem levantar exceção,
        garantindo que o fluxo de preços do chamador não é afetado.

        **Validates: Requirements 4.3, 10.1**
        """
        error = error_type(error_msg)
        worker = self._build_worker_with_failing_extractor(error)

        # Executar _process_intelligence — NÃO deve levantar exceção
        try:
            asyncio.run(
                worker._process_intelligence(
                    screenshot_bytes=screenshot,
                    competitor_id=competitor_id,
                    competitor_name="Concorrente Teste",
                    cycle_id=cycle_id,
                    home_url="https://example.com",
                )
            )
        except Exception as exc:
            pytest.fail(
                f"_process_intelligence propagou exceção "
                f"{type(exc).__name__}('{exc}') para o chamador. "
                f"Erro original: {error_type.__name__}('{error_msg}'). "
                f"Isso violaria o isolamento de falhas — preços "
                f"seriam impactados."
            )

    @settings(max_examples=100)
    @given(
        error_type=intelligence_error_type_strategy(),
        error_msg=error_message_strategy(),
        screenshot=screenshot_bytes_strategy(),
        competitor_id=st.uuids().map(str),
        cycle_id=st.uuids().map(str),
    )
    def test_property_8_price_store_nao_e_alterado_apos_falha(
        self,
        error_type: type,
        error_msg: str,
        screenshot: bytes,
        competitor_id: str,
        cycle_id: str,
    ) -> None:
        """Property 8: PriceStore não é chamado durante _process_intelligence.

        Para qualquer falha na extração de inteligência, o
        price_store (que contém PriceRecords) SHALL NÃO receber
        chamadas de save_record, delete, ou update — dados de
        preço permanecem intactos.

        **Validates: Requirements 4.3, 10.1**
        """
        from price_watchdog.worker.worker import Worker

        consumer = MagicMock()
        scraper = MagicMock()
        comparator = MagicMock()
        price_store = MagicMock()
        price_store.save_record = AsyncMock()
        price_store.delete_record = AsyncMock()
        screenshot_store = MagicMock()
        alert_service = MagicMock()

        # Intelligence extractor que falha
        intelligence_extractor = MagicMock()
        error = error_type(error_msg)
        intelligence_extractor.extract = AsyncMock(
            side_effect=error
        )

        # Intelligence store mockado
        intelligence_store = MagicMock()
        intelligence_store.save_record = AsyncMock()

        worker = Worker(
            consumer=consumer,
            scraper=scraper,
            comparator=comparator,
            price_store=price_store,
            screenshot_store=screenshot_store,
            alert_service=alert_service,
            intelligence_extractor=intelligence_extractor,
            intelligence_store=intelligence_store,
        )

        # Executar
        asyncio.run(
            worker._process_intelligence(
                screenshot_bytes=screenshot,
                competitor_id=competitor_id,
                competitor_name="Concorrente Teste",
                cycle_id=cycle_id,
                home_url="https://example.com",
            )
        )

        # price_store NÃO deve ter sido chamado de forma alguma
        assert not price_store.save_record.called, (
            f"price_store.save_record foi chamado durante "
            f"_process_intelligence com erro "
            f"{error_type.__name__}('{error_msg}'). "
            f"PriceRecords NÃO devem ser alterados por falhas "
            f"de inteligência."
        )

    @settings(max_examples=100)
    @given(
        error_type=intelligence_error_type_strategy(),
        error_msg=error_message_strategy(),
        competitor_id=st.uuids().map(str),
        cycle_id=st.uuids().map(str),
    )
    def test_property_8_falha_com_screenshot_none_nao_propaga(
        self,
        error_type: type,
        error_msg: str,
        competitor_id: str,
        cycle_id: str,
    ) -> None:
        """Property 8: Falha com screenshot_bytes=None não propaga exceção.

        Mesmo quando screenshot_bytes é None (cenário
        "screenshot_unavailable"), _process_intelligence SHALL
        completar sem levantar exceção, preservando o isolamento.

        **Validates: Requirements 4.3, 10.1**
        """
        error = error_type(error_msg)
        worker = self._build_worker_with_failing_extractor(error)

        # Executar com screenshot_bytes=None
        try:
            asyncio.run(
                worker._process_intelligence(
                    screenshot_bytes=None,
                    competitor_id=competitor_id,
                    competitor_name="Concorrente Teste",
                    cycle_id=cycle_id,
                    home_url="https://example.com",
                )
            )
        except Exception as exc:
            pytest.fail(
                f"_process_intelligence com screenshot=None "
                f"propagou exceção {type(exc).__name__}('{exc}'). "
                f"Deveria registrar 'screenshot_unavailable' sem "
                f"propagar erro."
            )

    @settings(max_examples=100)
    @given(
        error_type=intelligence_error_type_strategy(),
        error_msg=error_message_strategy(),
        screenshot=screenshot_bytes_strategy(),
        competitor_id=st.uuids().map(str),
        cycle_id=st.uuids().map(str),
    )
    def test_property_8_metodo_retorna_none_apos_falha(
        self,
        error_type: type,
        error_msg: str,
        screenshot: bytes,
        competitor_id: str,
        cycle_id: str,
    ) -> None:
        """Property 8: _process_intelligence retorna None após falha.

        Para qualquer falha na extração de inteligência, o método
        SHALL completar normalmente e retornar None (método async
        void), sem alterar o fluxo do chamador.

        **Validates: Requirements 4.3, 10.1**
        """
        error = error_type(error_msg)
        worker = self._build_worker_with_failing_extractor(error)

        # Executar e verificar retorno
        result = asyncio.run(
            worker._process_intelligence(
                screenshot_bytes=screenshot,
                competitor_id=competitor_id,
                competitor_name="Concorrente Teste",
                cycle_id=cycle_id,
                home_url="https://example.com",
            )
        )

        assert result is None, (
            f"_process_intelligence deveria retornar None "
            f"após falha {error_type.__name__}('{error_msg}'), "
            f"mas retornou {result!r}."
        )


# --- Estratégias (generators) para Property 7 ---


def intelligence_flag_strategy() -> st.SearchStrategy[bool]:
    """Gera valores booleanos para intelligence_enabled."""
    return st.booleans()


def base_url_strategy() -> st.SearchStrategy[str]:
    """Gera URLs base válidas para concorrentes."""
    return st.sampled_from([
        "https://www.claro.com.br",
        "https://www.vivo.com.br",
        "https://www.oi.com.br",
        "https://www.tim.com.br",
        "https://www.sky.com.br",
        "https://www.net.com.br",
    ])


def intelligence_home_url_strategy() -> st.SearchStrategy:
    """Gera intelligence_home_url (pode ser None ou URL válida)."""
    return st.one_of(
        st.none(),
        st.sampled_from([
            "https://www.claro.com.br/tv",
            "https://www.vivo.com.br/planos",
            "https://assine.oi.com.br",
            "https://www.tim.com.br/ofertas",
        ]),
    )


def competitor_with_flags_strategy() -> st.SearchStrategy[dict]:
    """Gera dados de concorrente com flags de inteligência variados.

    Retorna um dict com os dados necessários para montar um
    MagicMock de Competitor com intelligence_enabled, base_url
    e intelligence_home_url.
    """
    return st.fixed_dictionaries({
        "id": st.uuids(),
        "name": st.sampled_from([
            "Claro", "Vivo", "Oi", "Tim", "Sky", "Net",
        ]),
        "base_url": base_url_strategy(),
        "intelligence_enabled": intelligence_flag_strategy(),
        "intelligence_home_url": intelligence_home_url_strategy(),
        "is_active": st.just(True),
    })


def competitor_set_strategy() -> st.SearchStrategy[list[dict]]:
    """Gera conjuntos de 2-8 concorrentes com flags variados.

    Garante que pelo menos um tem intelligence_enabled=True
    e pelo menos um tem intelligence_enabled=False para testar
    a filtragem corretamente.
    """
    return st.lists(
        competitor_with_flags_strategy(),
        min_size=2,
        max_size=8,
    ).filter(
        lambda comps: (
            any(c["intelligence_enabled"] for c in comps)
            and any(not c["intelligence_enabled"] for c in comps)
        )
    )


def mixed_competitor_set_strategy() -> st.SearchStrategy[list[dict]]:
    """Gera conjuntos de 1-10 concorrentes (sem filtro obrigatório).

    Pode ter todos com True, todos com False, ou mistura.
    """
    return st.lists(
        competitor_with_flags_strategy(),
        min_size=1,
        max_size=10,
    )


# --- Property 7 Tests ---


@pytest.mark.property
class TestIntelligenceEnabledFilteringProperty:
    """Property 7: Filtragem por intelligence_enabled.

    Para qualquer conjunto de Competitors onde um subconjunto tem
    intelligence_enabled=true e o restante intelligence_enabled=false,
    o sistema SHALL incluir na extração de inteligência APENAS os
    concorrentes com flag=true, e SHALL excluir os demais sem remover
    dados históricos existentes.

    Feature: competitor-intelligence-expansion, Property 7

    **Validates: Requirements 4.1, 8.2, 8.3**
    """

    def _build_mock_configs(
        self, competitors_data: list[dict]
    ) -> list:
        """Constrói lista de ProductConfig mocks a partir dos dados.

        Cada concorrente gera 1 ProductConfig com
        extraction_strategy="ai_all" para testar o agrupamento
        do coordinator._publish_tasks.
        """
        configs = []
        for comp_data in competitors_data:
            config = MagicMock()
            config.id = comp_data["id"]
            config.competitor_id = comp_data["id"]
            config.product_name = "Plano Base"
            config.page_url = f"{comp_data['base_url']}/planos"
            config.extraction_strategy = "ai_all"
            config.selector_or_pattern = ""
            config.our_price = 99.90

            # Mock do competitor associado
            competitor = MagicMock()
            competitor.name = comp_data["name"]
            competitor.base_url = comp_data["base_url"]
            competitor.intelligence_enabled = (
                comp_data["intelligence_enabled"]
            )
            competitor.intelligence_home_url = (
                comp_data["intelligence_home_url"]
            )

            config.competitor = competitor
            configs.append(config)

        return configs

    def _run_publish_tasks(
        self, competitors_data: list[dict]
    ) -> list:
        """Executa _publish_tasks e retorna as mensagens criadas.

        Monta mocks do coordinator e executa _publish_tasks
        capturando as mensagens passadas ao publisher.
        """
        from price_watchdog.coordinator.coordinator import (
            PriceMonitoringCoordinator,
        )

        configs = self._build_mock_configs(competitors_data)

        # Mock do publisher que captura mensagens
        publisher = MagicMock()
        captured_messages = []

        async def capture_publish(messages, batch_size=10):
            captured_messages.extend(messages)
            return len(messages)

        publisher.publish_all = AsyncMock(
            side_effect=capture_publish
        )

        # Mocks restantes
        consolidator = MagicMock()
        price_store = MagicMock()
        competitor_manager = MagicMock()

        coordinator = PriceMonitoringCoordinator(
            publisher=publisher,
            consolidator=consolidator,
            price_store=price_store,
            competitor_manager=competitor_manager,
        )

        # Mock do cycle
        cycle = MagicMock()
        cycle.id = "cycle-test-id"

        asyncio.run(coordinator._publish_tasks(cycle, configs))

        return captured_messages

    @settings(max_examples=100)
    @given(competitors=competitor_set_strategy())
    def test_property_7_mensagens_com_intel_true_quando_flag_true(
        self, competitors: list[dict]
    ) -> None:
        """Property 7: Competitors com intelligence_enabled=True geram
        mensagens com intelligence_enabled=True.

        Para qualquer conjunto de Competitors, as mensagens geradas
        para aqueles com flag=True SHALL ter intelligence_enabled=True
        no PriceCheckMessage.

        **Validates: Requirements 4.1, 8.2, 8.3**
        """
        messages = self._run_publish_tasks(competitors)

        # Construir mapa competitor_id -> intelligence_enabled
        comp_map = {
            str(c["id"]): c["intelligence_enabled"]
            for c in competitors
        }

        for msg in messages:
            expected_flag = comp_map.get(msg.competitor_id, False)
            if expected_flag:
                assert msg.intelligence_enabled is True, (
                    f"Competitor '{msg.competitor_name}' tem "
                    f"intelligence_enabled=True mas a mensagem "
                    f"tem intelligence_enabled={msg.intelligence_enabled}."
                )

    @settings(max_examples=100)
    @given(competitors=competitor_set_strategy())
    def test_property_7_mensagens_com_intel_false_quando_flag_false(
        self, competitors: list[dict]
    ) -> None:
        """Property 7: Competitors com intelligence_enabled=False geram
        mensagens com intelligence_enabled=False.

        Para qualquer conjunto de Competitors, as mensagens geradas
        para aqueles com flag=False SHALL ter intelligence_enabled=False
        no PriceCheckMessage, excluindo-os da extração de inteligência.

        **Validates: Requirements 4.1, 8.2, 8.3**
        """
        messages = self._run_publish_tasks(competitors)

        # Construir mapa competitor_id -> intelligence_enabled
        comp_map = {
            str(c["id"]): c["intelligence_enabled"]
            for c in competitors
        }

        for msg in messages:
            expected_flag = comp_map.get(msg.competitor_id, False)
            if not expected_flag:
                assert msg.intelligence_enabled is False, (
                    f"Competitor '{msg.competitor_name}' tem "
                    f"intelligence_enabled=False mas a mensagem "
                    f"tem intelligence_enabled={msg.intelligence_enabled}."
                )

    @settings(max_examples=100)
    @given(competitors=competitor_set_strategy())
    def test_property_7_intel_home_url_definida_apenas_quando_enabled(
        self, competitors: list[dict]
    ) -> None:
        """Property 7: intelligence_home_url definida somente quando
        intelligence_enabled=True, com fallback para base_url.

        Para qualquer conjunto de Competitors:
        - Se intelligence_enabled=True E intelligence_home_url está
          configurada → mensagem usa intelligence_home_url
        - Se intelligence_enabled=True E intelligence_home_url é None →
          mensagem usa base_url (fallback)
        - Se intelligence_enabled=False → mensagem tem
          intelligence_home_url=None

        **Validates: Requirements 4.1, 8.2, 8.3**
        """
        messages = self._run_publish_tasks(competitors)

        # Construir mapa competitor_id -> dados completos
        comp_map = {str(c["id"]): c for c in competitors}

        for msg in messages:
            comp_data = comp_map.get(msg.competitor_id)
            if comp_data is None:
                continue

            if comp_data["intelligence_enabled"]:
                # Quando enabled=True, home_url deve estar definida
                expected_url = (
                    comp_data["intelligence_home_url"]
                    or comp_data["base_url"]
                )
                assert msg.intelligence_home_url == expected_url, (
                    f"Competitor '{msg.competitor_name}' com "
                    f"intelligence_enabled=True deveria ter "
                    f"intelligence_home_url='{expected_url}', "
                    f"mas tem '{msg.intelligence_home_url}'. "
                    f"(home_url={comp_data['intelligence_home_url']}, "
                    f"base_url={comp_data['base_url']})"
                )
                assert msg.intelligence_home_url is not None, (
                    f"Competitor '{msg.competitor_name}' com "
                    f"intelligence_enabled=True deveria ter "
                    f"intelligence_home_url definida (não None). "
                    f"Fallback para base_url deveria garantir isso."
                )
            else:
                # Quando enabled=False, home_url deve ser None
                assert msg.intelligence_home_url is None, (
                    f"Competitor '{msg.competitor_name}' com "
                    f"intelligence_enabled=False deveria ter "
                    f"intelligence_home_url=None, mas tem "
                    f"'{msg.intelligence_home_url}'."
                )

    @settings(max_examples=100)
    @given(competitors=mixed_competitor_set_strategy())
    def test_property_7_todos_competitors_geram_mensagens(
        self, competitors: list[dict]
    ) -> None:
        """Property 7: Todos os competitors geram mensagens (flag
        não impede geração de mensagem, apenas define o campo).

        A filtragem por intelligence_enabled NÃO remove
        competitors da lista de mensagens — apenas define
        o campo intelligence_enabled na mensagem. Isso garante
        que dados históricos de preços não são impactados.

        **Validates: Requirements 4.1, 8.2, 8.3**
        """
        messages = self._run_publish_tasks(competitors)

        # Como todos usam ai_all, deve agrupar por competitor_id
        # Cada competitor_id único gera 1 mensagem
        unique_competitor_ids = {
            str(c["id"]) for c in competitors
        }
        message_competitor_ids = {
            msg.competitor_id for msg in messages
        }

        assert unique_competitor_ids == message_competitor_ids, (
            f"Todos os competitors deveriam gerar mensagens. "
            f"Esperado: {len(unique_competitor_ids)} mensagens, "
            f"Obtido: {len(message_competitor_ids)}. "
            f"Faltantes: {unique_competitor_ids - message_competitor_ids}"
        )

# --- Estratégias (generators) para Property 12 ---


def keyword_strategy() -> st.SearchStrategy[str]:
    """Gera palavras-chave comerciais para testes de comunicação.

    Produz strings de 1-50 caracteres com letras, números e
    espaços, simulando keywords comerciais reais.
    """
    return st.text(
        alphabet=st.characters(
            whitelist_categories=("L", "N", "Z"),
            blacklist_characters="\x00\n\r",
        ),
        min_size=1,
        max_size=50,
    ).filter(lambda s: s.strip() != "")


def keyword_list_strategy(
    min_size: int = 3, max_size: int = 15
) -> st.SearchStrategy[list[str]]:
    """Gera listas de keywords de tamanho variável.

    Args:
        min_size: Mínimo de keywords na lista.
        max_size: Máximo de keywords na lista.

    Returns:
        Estratégia que produz listas de keywords únicas.
    """
    return st.lists(
        keyword_strategy(),
        min_size=min_size,
        max_size=max_size,
    )


def banner_text_strategy(
    min_size: int = 10, max_size: int = 500
) -> st.SearchStrategy[str]:
    """Gera textos de banner com tamanho variável.

    Produz textos que simulam descrições de banners comerciais
    com letras, números, pontuação e espaços.
    """
    return st.text(
        alphabet=st.characters(
            whitelist_categories=("L", "N", "Z", "P"),
            blacklist_characters="\x00",
        ),
        min_size=min_size,
        max_size=max_size,
    ).filter(lambda s: s.strip() != "")


def communication_record_mock(
    keywords,
    banner,
) -> MagicMock:
    """Cria MagicMock de CompetitorIntelligenceRecord com dados de comunicação.

    Args:
        keywords: Lista de keywords comerciais (ou None).
        banner: Descrição do banner (ou None).

    Returns:
        MagicMock com spec de CompetitorIntelligenceRecord.
    """
    from price_watchdog.models.intelligence_entities import (
        CompetitorIntelligenceRecord,
    )

    record = MagicMock(spec=CompetitorIntelligenceRecord)
    record.commercial_keywords = keywords
    record.home_banner_description = banner
    return record


# --- Property 12 Tests ---


@pytest.mark.property
class TestCommunicationChangeDetectionProperty:
    """Property 12: Detecção de mudanças significativas em comunicação comercial.

    Para quaisquer dois conjuntos de keywords e descrições de banner,
    o sistema SHALL gerar alerta "communication_change" se e somente se
    mais de 50% das keywords mudaram (interseção/total < 0.5) OU a
    similaridade textual do banner for inferior a 60%.

    Feature: competitor-intelligence-expansion, Property 12

    **Validates: Requirements 7.4**
    """

    @settings(max_examples=100)
    @given(keywords=keyword_list_strategy())
    def test_property_12_keywords_identicas_change_pct_zero(
        self, keywords: list[str]
    ) -> None:
        """Property 12: Keywords idênticas → change_pct = 0.0 (sem alerta).

        Quando current_keywords e previous_keywords são idênticos,
        _calculate_keyword_change_pct SHALL retornar 0.0, indicando
        nenhuma mudança e portanto nenhum alerta de keywords.

        **Validates: Requirements 7.4**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()

        change_pct = detector._calculate_keyword_change_pct(
            keywords, keywords
        )

        assert change_pct == 0.0, (
            f"Keywords idênticas deveriam ter change_pct=0.0, "
            f"mas obteve {change_pct}. Keywords: {keywords[:3]}..."
        )

    @settings(max_examples=100)
    @given(
        current=keyword_list_strategy(),
        previous=keyword_list_strategy(),
    )
    def test_property_12_keywords_completamente_diferentes_change_pct_1(
        self, current: list[str], previous: list[str]
    ) -> None:
        """Property 12: Keywords sem interseção → change_pct = 1.0 (alerta).

        Quando current_keywords e previous_keywords não possuem
        nenhuma interseção (após normalização lowercase/strip),
        _calculate_keyword_change_pct SHALL retornar 1.0.

        **Validates: Requirements 7.4**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        # Garantir que não há interseção
        current_set = set(kw.lower().strip() for kw in current)
        previous_set = set(kw.lower().strip() for kw in previous)
        assume(len(current_set & previous_set) == 0)
        assume(len(current_set | previous_set) > 0)

        detector = ChangeDetector()

        change_pct = detector._calculate_keyword_change_pct(
            current, previous
        )

        assert change_pct == 1.0, (
            f"Keywords completamente diferentes deveriam ter "
            f"change_pct=1.0, mas obteve {change_pct}. "
            f"Current: {current[:3]}, Previous: {previous[:3]}"
        )

    @settings(max_examples=100)
    @given(
        current=keyword_list_strategy(),
        previous=keyword_list_strategy(),
    )
    def test_property_12_keyword_change_pct_gera_alerta_quando_maior_05(
        self, current: list[str], previous: list[str]
    ) -> None:
        """Property 12: change_pct > 0.5 → alerta "communication_change" gerado.

        Quando _calculate_keyword_change_pct retorna valor > 0.5,
        _compare_communication SHALL gerar alerta com
        alert_type="communication_change" e
        attribute_name="commercial_keywords".

        **Validates: Requirements 7.4**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()

        change_pct = detector._calculate_keyword_change_pct(
            current, previous
        )
        assume(change_pct > 0.5)

        # Montar records mockados com keywords e banner idêntico
        # (para isolar o teste de keywords)
        same_banner = "Banner de teste com oferta especial"
        current_record = communication_record_mock(
            keywords=current, banner=same_banner
        )
        previous_record = communication_record_mock(
            keywords=previous, banner=same_banner
        )

        alerts = detector._compare_communication(
            current_record, previous_record
        )

        # Deve existir ao menos um alerta de keywords
        keyword_alerts = [
            a for a in alerts
            if a.attribute_name == "commercial_keywords"
        ]
        assert len(keyword_alerts) >= 1, (
            f"change_pct={change_pct:.3f} (> 0.5) deveria gerar "
            f"alerta de keywords, mas nenhum foi gerado. "
            f"Current: {current[:3]}, Previous: {previous[:3]}"
        )
        # Verificar tipo do alerta
        assert keyword_alerts[0].alert_type == "communication_change", (
            f"Alerta de keywords deveria ter "
            f"alert_type='communication_change', mas tem "
            f"'{keyword_alerts[0].alert_type}'."
        )

    @settings(max_examples=100)
    @given(
        current=keyword_list_strategy(),
        previous=keyword_list_strategy(),
    )
    def test_property_12_keyword_change_pct_nao_gera_alerta_quando_menor_igual_05(
        self, current: list[str], previous: list[str]
    ) -> None:
        """Property 12: change_pct ≤ 0.5 → sem alerta de keywords.

        Quando _calculate_keyword_change_pct retorna valor ≤ 0.5,
        _compare_communication SHALL NÃO gerar alerta com
        attribute_name="commercial_keywords".

        **Validates: Requirements 7.4**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()

        change_pct = detector._calculate_keyword_change_pct(
            current, previous
        )
        assume(change_pct <= 0.5)

        # Montar records com keywords e banner idêntico
        same_banner = "Banner inalterado para teste"
        current_record = communication_record_mock(
            keywords=current, banner=same_banner
        )
        previous_record = communication_record_mock(
            keywords=previous, banner=same_banner
        )

        alerts = detector._compare_communication(
            current_record, previous_record
        )

        # NÃO deve existir alerta de keywords
        keyword_alerts = [
            a for a in alerts
            if a.attribute_name == "commercial_keywords"
        ]
        assert len(keyword_alerts) == 0, (
            f"change_pct={change_pct:.3f} (≤ 0.5) NÃO deveria gerar "
            f"alerta de keywords, mas gerou {len(keyword_alerts)}. "
            f"Current: {current[:3]}, Previous: {previous[:3]}"
        )

    @settings(max_examples=100)
    @given(
        current=keyword_list_strategy(),
        previous=keyword_list_strategy(),
    )
    def test_property_12_keyword_change_pct_sempre_entre_0_e_1(
        self, current: list[str], previous: list[str]
    ) -> None:
        """Property 12: _calculate_keyword_change_pct sempre retorna [0.0, 1.0].

        Para quaisquer duas listas de keywords, o resultado
        SHALL ser um float entre 0.0 e 1.0 inclusive.

        **Validates: Requirements 7.4**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()

        change_pct = detector._calculate_keyword_change_pct(
            current, previous
        )

        assert 0.0 <= change_pct <= 1.0, (
            f"change_pct deveria estar em [0.0, 1.0], "
            f"mas obteve {change_pct}. "
            f"Current ({len(current)}): {current[:3]}, "
            f"Previous ({len(previous)}): {previous[:3]}"
        )

    @settings(max_examples=100)
    @given(
        text_a=banner_text_strategy(),
        text_b=banner_text_strategy(),
    )
    def test_property_12_text_similarity_sempre_entre_0_e_1(
        self, text_a: str, text_b: str
    ) -> None:
        """Property 12: _calculate_text_similarity sempre retorna [0.0, 1.0].

        Para quaisquer dois textos, o resultado SHALL ser um float
        entre 0.0 e 1.0 inclusive.

        **Validates: Requirements 7.4**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()

        similarity = detector._calculate_text_similarity(
            text_a, text_b
        )

        assert 0.0 <= similarity <= 1.0, (
            f"Similaridade deveria estar em [0.0, 1.0], "
            f"mas obteve {similarity}. "
            f"text_a (len={len(text_a)}): '{text_a[:30]}...', "
            f"text_b (len={len(text_b)}): '{text_b[:30]}...'"
        )

    @settings(max_examples=100)
    @given(text=banner_text_strategy())
    def test_property_12_text_similarity_identica_retorna_1(
        self, text: str
    ) -> None:
        """Property 12: Textos idênticos → similaridade = 1.0.

        Quando text_a == text_b, _calculate_text_similarity SHALL
        retornar 1.0. Similaridade perfeita indica banners idênticos
        e portanto nenhum alerta de banner.

        **Validates: Requirements 7.4**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()

        similarity = detector._calculate_text_similarity(text, text)

        assert similarity == 1.0, (
            f"Textos idênticos deveriam ter similaridade=1.0, "
            f"mas obteve {similarity}. "
            f"Texto: '{text[:50]}...'"
        )

    @settings(max_examples=100)
    @given(
        current_banner=banner_text_strategy(),
        previous_banner=banner_text_strategy(),
    )
    def test_property_12_banner_baixa_similaridade_gera_alerta(
        self, current_banner: str, previous_banner: str
    ) -> None:
        """Property 12: Banner com similaridade < 0.6 → alerta gerado.

        Quando _calculate_text_similarity retorna valor < 0.6,
        _compare_communication SHALL gerar alerta com
        alert_type="communication_change" e
        attribute_name="home_banner_description".

        **Validates: Requirements 7.4**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()

        similarity = detector._calculate_text_similarity(
            current_banner, previous_banner
        )
        assume(similarity < 0.6)

        # Montar records com keywords idênticas (isolar teste de banner)
        same_keywords = ["oferta", "desconto", "streaming"]
        current_record = communication_record_mock(
            keywords=same_keywords, banner=current_banner
        )
        previous_record = communication_record_mock(
            keywords=same_keywords, banner=previous_banner
        )

        alerts = detector._compare_communication(
            current_record, previous_record
        )

        # Deve existir alerta de banner
        banner_alerts = [
            a for a in alerts
            if a.attribute_name == "home_banner_description"
        ]
        assert len(banner_alerts) >= 1, (
            f"Similaridade={similarity:.3f} (< 0.6) deveria gerar "
            f"alerta de banner, mas nenhum foi gerado. "
            f"Current: '{current_banner[:30]}...', "
            f"Previous: '{previous_banner[:30]}...'"
        )
        assert banner_alerts[0].alert_type == "communication_change", (
            f"Alerta de banner deveria ter "
            f"alert_type='communication_change', mas tem "
            f"'{banner_alerts[0].alert_type}'."
        )

    @settings(max_examples=100)
    @given(
        base_banner=banner_text_strategy(min_size=20, max_size=300),
        suffix=st.text(
            alphabet=st.characters(
                whitelist_categories=("L", "N", "Z"),
                blacklist_characters="\x00",
            ),
            min_size=0,
            max_size=30,
        ),
    )
    def test_property_12_banner_alta_similaridade_nao_gera_alerta(
        self, base_banner: str, suffix: str
    ) -> None:
        """Property 12: Banner com similaridade ≥ 0.6 → sem alerta de banner.

        Quando _calculate_text_similarity retorna valor ≥ 0.6,
        _compare_communication SHALL NÃO gerar alerta com
        attribute_name="home_banner_description".

        Para gerar pares com alta similaridade, usamos um texto
        base e apenas adicionamos um sufixo curto ao segundo.

        **Validates: Requirements 7.4**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()

        # Gerar banner similar ao base adicionando sufixo curto
        current_banner = base_banner
        previous_banner = base_banner + suffix

        similarity = detector._calculate_text_similarity(
            current_banner, previous_banner
        )
        assume(similarity >= 0.6)

        # Montar records com keywords idênticas (isolar teste de banner)
        same_keywords = ["oferta", "desconto", "streaming"]
        current_record = communication_record_mock(
            keywords=same_keywords, banner=current_banner
        )
        previous_record = communication_record_mock(
            keywords=same_keywords, banner=previous_banner
        )

        alerts = detector._compare_communication(
            current_record, previous_record
        )

        # NÃO deve existir alerta de banner
        banner_alerts = [
            a for a in alerts
            if a.attribute_name == "home_banner_description"
        ]
        assert len(banner_alerts) == 0, (
            f"Similaridade={similarity:.3f} (≥ 0.6) NÃO deveria "
            f"gerar alerta de banner, mas gerou "
            f"{len(banner_alerts)}. "
            f"Current: '{current_banner[:30]}...', "
            f"Previous: '{previous_banner[:30]}...'"
        )

# --- Estratégias (generators) para Property 11 ---


def _composition_attribute_value(
    attr_name: str,
) -> st.SearchStrategy:
    """Gera valor válido para um atributo de PackageComposition.

    Retorna estratégia apropriada para cada tipo de atributo:
    - Float para preços
    - Int para canais, telas, velocidades, período
    - Bool para has_fiber, has_mobile_internet
    - String para streamings
    """
    if attr_name in ("default_price", "promotional_price"):
        return st.floats(
            min_value=10.0,
            max_value=5000.0,
            allow_nan=False,
            allow_infinity=False,
        )
    elif attr_name == "promotional_period_months":
        return st.integers(min_value=1, max_value=36)
    elif attr_name in (
        "linear_channels",
        "simultaneous_screens",
        "fiber_speed_mbps",
        "mobile_speed_mbps",
    ):
        return st.integers(min_value=0, max_value=9999)
    elif attr_name in ("has_fiber", "has_mobile_internet"):
        return st.booleans()
    elif attr_name in (
        "bundled_streaming_1",
        "bundled_streaming_2",
        "bundled_streaming_3",
    ):
        return st.sampled_from([
            "Netflix", "Disney+", "Paramount+",
            "Globoplay", "HBO Max", "Star+",
            None,
        ])
    else:
        return st.none()


# Atributos comparáveis de PackageComposition (conforme change_detector)
_COMPOSITION_ATTRS_FOR_COMPARISON = [
    "default_price",
    "promotional_price",
    "promotional_period_months",
    "linear_channels",
    "simultaneous_screens",
    "has_fiber",
    "fiber_speed_mbps",
    "has_mobile_internet",
    "mobile_speed_mbps",
    "bundled_streaming_1",
    "bundled_streaming_2",
    "bundled_streaming_3",
]


def _make_package_mock(
    plan_name: str, attrs: dict
) -> MagicMock:
    """Cria um MagicMock de PackageComposition com atributos dados.

    Args:
        plan_name: Nome do plano para o pacote.
        attrs: Dicionário de atributo → valor para setar no mock.

    Returns:
        MagicMock com plan_name e atributos setados via getattr.
    """
    from price_watchdog.models.intelligence_entities import (
        PackageComposition,
    )

    mock = MagicMock(spec=PackageComposition)
    mock.plan_name = plan_name

    for attr_name in _COMPOSITION_ATTRS_FOR_COMPARISON:
        setattr(mock, attr_name, attrs.get(attr_name))

    return mock


def identical_composition_pair_strategy() -> st.SearchStrategy:
    """Gera par de composições idênticas (mesmos atributos).

    Usado para verificar que composições iguais → 0 alertas.
    """
    return st.fixed_dictionaries({
        "plan_name": st.text(
            alphabet=st.characters(
                whitelist_categories=("L", "N", "Z"),
            ),
            min_size=1,
            max_size=30,
        ).filter(lambda s: s.strip() != ""),
        "default_price": st.floats(
            min_value=10.0,
            max_value=5000.0,
            allow_nan=False,
            allow_infinity=False,
        ),
        "promotional_price": st.one_of(
            st.none(),
            st.floats(
                min_value=10.0,
                max_value=4999.0,
                allow_nan=False,
                allow_infinity=False,
            ),
        ),
        "promotional_period_months": st.one_of(
            st.none(),
            st.integers(min_value=1, max_value=36),
        ),
        "linear_channels": st.one_of(
            st.none(),
            st.integers(min_value=0, max_value=500),
        ),
        "simultaneous_screens": st.one_of(
            st.none(),
            st.integers(min_value=0, max_value=10),
        ),
        "has_fiber": st.one_of(st.none(), st.booleans()),
        "fiber_speed_mbps": st.one_of(
            st.none(),
            st.integers(min_value=0, max_value=2000),
        ),
        "has_mobile_internet": st.one_of(st.none(), st.booleans()),
        "mobile_speed_mbps": st.one_of(
            st.none(),
            st.integers(min_value=0, max_value=1000),
        ),
        "bundled_streaming_1": st.one_of(
            st.none(),
            st.sampled_from(["Netflix", "Disney+", "Paramount+"]),
        ),
        "bundled_streaming_2": st.one_of(
            st.none(),
            st.sampled_from(["Globoplay", "HBO Max", "Star+"]),
        ),
        "bundled_streaming_3": st.one_of(
            st.none(),
            st.sampled_from(["Apple TV+", "Amazon Prime Video"]),
        ),
    })


def changed_attributes_strategy() -> st.SearchStrategy:
    """Gera subconjunto não-vazio de atributos a serem alterados.

    Retorna frozenset com 1 a 5 atributos a mudar entre
    current e previous.
    """
    return st.frozensets(
        st.sampled_from(_COMPOSITION_ATTRS_FOR_COMPARISON),
        min_size=1,
        max_size=5,
    )


def _different_value_for_attr(
    attr_name: str, original_value
) -> st.SearchStrategy:
    """Gera um valor diferente do original para o atributo dado.

    Garante que o valor gerado é diferente de original_value
    para que a mudança seja detectável.
    """
    if attr_name in ("default_price", "promotional_price"):
        if original_value is None:
            return st.just(99.99)
        return st.floats(
            min_value=10.0,
            max_value=5000.0,
            allow_nan=False,
            allow_infinity=False,
        ).filter(lambda v: v != original_value)
    elif attr_name == "promotional_period_months":
        if original_value is None:
            return st.just(12)
        return st.integers(
            min_value=1, max_value=36
        ).filter(lambda v: v != original_value)
    elif attr_name in (
        "linear_channels",
        "simultaneous_screens",
        "fiber_speed_mbps",
        "mobile_speed_mbps",
    ):
        if original_value is None:
            return st.just(100)
        return st.integers(
            min_value=0, max_value=9999
        ).filter(lambda v: v != original_value)
    elif attr_name in ("has_fiber", "has_mobile_internet"):
        if original_value is None:
            return st.just(True)
        return st.just(not original_value)
    elif attr_name in (
        "bundled_streaming_1",
        "bundled_streaming_2",
        "bundled_streaming_3",
    ):
        all_options = [
            "Netflix", "Disney+", "Paramount+",
            "Globoplay", "HBO Max", "Star+",
        ]
        if original_value is None:
            return st.just("Netflix")
        return st.sampled_from(
            [s for s in all_options if s != original_value]
        )
    else:
        return st.none()


def new_package_strategy() -> st.SearchStrategy[dict]:
    """Gera dados para um pacote novo (sem equivalente anterior).

    Pacotes novos devem gerar alertas para cada atributo não-null.
    """
    return st.fixed_dictionaries({
        "plan_name": st.text(
            alphabet=st.characters(
                whitelist_categories=("L", "N"),
            ),
            min_size=3,
            max_size=20,
        ).filter(lambda s: s.strip() != ""),
        "default_price": st.one_of(
            st.none(),
            st.floats(
                min_value=10.0,
                max_value=5000.0,
                allow_nan=False,
                allow_infinity=False,
            ),
        ),
        "promotional_price": st.one_of(
            st.none(),
            st.floats(
                min_value=10.0,
                max_value=4999.0,
                allow_nan=False,
                allow_infinity=False,
            ),
        ),
        "linear_channels": st.one_of(
            st.none(),
            st.integers(min_value=0, max_value=500),
        ),
        "simultaneous_screens": st.one_of(
            st.none(),
            st.integers(min_value=1, max_value=10),
        ),
        "has_fiber": st.one_of(st.none(), st.booleans()),
        "fiber_speed_mbps": st.one_of(
            st.none(),
            st.integers(min_value=100, max_value=2000),
        ),
        "has_mobile_internet": st.one_of(st.none(), st.booleans()),
        "mobile_speed_mbps": st.one_of(
            st.none(),
            st.integers(min_value=50, max_value=1000),
        ),
        "bundled_streaming_1": st.one_of(
            st.none(),
            st.sampled_from(["Netflix", "Disney+"]),
        ),
        "bundled_streaming_2": st.one_of(
            st.none(),
            st.sampled_from(["Globoplay", "HBO Max"]),
        ),
        "bundled_streaming_3": st.one_of(st.none(), st.just("Star+")),
        "promotional_period_months": st.one_of(
            st.none(),
            st.integers(min_value=1, max_value=36),
        ),
    })


# --- Property 11 Tests ---


@pytest.mark.property
class TestCompositionChangeDetectionProperty:
    """Property 11: Detecção de mudanças em composição de pacotes.

    Para qualquer par de Competitor_Intelligence_Records consecutivos
    (anterior com sucesso e atual com sucesso) do mesmo concorrente,
    o ChangeDetector SHALL identificar corretamente todas as diferenças
    em atributos de composição (preço, canais, telas, streamings,
    velocidades) e gerar um alerta "package_composition_change" para
    cada atributo alterado, contendo valor anterior e valor atual.

    Feature: competitor-intelligence-expansion, Property 11

    **Validates: Requirements 7.1, 7.3**
    """

    @settings(max_examples=100)
    @given(comp_data=identical_composition_pair_strategy())
    def test_property_11_composicoes_identicas_zero_alertas(
        self, comp_data: dict
    ) -> None:
        """Property 11: Composições idênticas → nenhum alerta.

        Para qualquer par de pacotes com todos os atributos iguais
        (current == previous), _compare_compositions SHALL retornar
        lista vazia de alertas.

        **Validates: Requirements 7.1, 7.3**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()
        plan_name = comp_data["plan_name"]
        attrs = {
            k: v for k, v in comp_data.items() if k != "plan_name"
        }

        current_pkg = _make_package_mock(plan_name, attrs)
        previous_pkg = _make_package_mock(plan_name, attrs)

        alerts = detector._compare_compositions(
            current=[current_pkg],
            previous=[previous_pkg],
        )

        assert len(alerts) == 0, (
            f"Composições idênticas deveriam gerar 0 alertas, "
            f"mas geraram {len(alerts)}. "
            f"Plan: '{plan_name}', Attrs: {attrs}"
        )

    @settings(max_examples=100)
    @given(
        comp_data=identical_composition_pair_strategy(),
        attrs_to_change=changed_attributes_strategy(),
    )
    def test_property_11_atributos_alterados_geram_alertas(
        self, comp_data: dict, attrs_to_change: frozenset
    ) -> None:
        """Property 11: Atributos alterados → 1 alerta por atributo.

        Para qualquer par de pacotes onde N atributos diferem entre
        current e previous, _compare_compositions SHALL retornar
        exatamente N alertas, cada um do tipo
        "package_composition_change" com o nome do atributo alterado.

        **Validates: Requirements 7.1, 7.3**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()
        plan_name = comp_data["plan_name"]

        # Construir atributos do registro anterior (base)
        previous_attrs = {
            k: v for k, v in comp_data.items() if k != "plan_name"
        }

        # Construir atributos do registro atual (com mudanças)
        current_attrs = dict(previous_attrs)
        for attr_name in attrs_to_change:
            original_val = previous_attrs.get(attr_name)
            # Gerar valor diferente de forma determinística
            if attr_name in ("default_price", "promotional_price"):
                if original_val is None:
                    current_attrs[attr_name] = 199.99
                else:
                    current_attrs[attr_name] = original_val + 50.0
            elif attr_name == "promotional_period_months":
                if original_val is None:
                    current_attrs[attr_name] = 6
                else:
                    current_attrs[attr_name] = (
                        (original_val % 36) + 1
                    )
            elif attr_name in (
                "linear_channels",
                "simultaneous_screens",
                "fiber_speed_mbps",
                "mobile_speed_mbps",
            ):
                if original_val is None:
                    current_attrs[attr_name] = 100
                else:
                    current_attrs[attr_name] = original_val + 10
            elif attr_name in ("has_fiber", "has_mobile_internet"):
                if original_val is None:
                    current_attrs[attr_name] = True
                else:
                    current_attrs[attr_name] = not original_val
            elif attr_name in (
                "bundled_streaming_1",
                "bundled_streaming_2",
                "bundled_streaming_3",
            ):
                if original_val is None:
                    current_attrs[attr_name] = "Netflix"
                elif original_val == "Netflix":
                    current_attrs[attr_name] = "Disney+"
                else:
                    current_attrs[attr_name] = "Netflix"

        previous_pkg = _make_package_mock(plan_name, previous_attrs)
        current_pkg = _make_package_mock(plan_name, current_attrs)

        alerts = detector._compare_compositions(
            current=[current_pkg],
            previous=[previous_pkg],
        )

        # Verificar número de alertas
        expected_count = len(attrs_to_change)
        assert len(alerts) == expected_count, (
            f"Deveria gerar exatamente {expected_count} alertas "
            f"(1 por atributo alterado), mas gerou {len(alerts)}. "
            f"Atributos alterados: {attrs_to_change}. "
            f"Plan: '{plan_name}'"
        )

        # Verificar tipo de alerta
        for alert in alerts:
            assert alert.alert_type == "package_composition_change", (
                f"Tipo do alerta deveria ser "
                f"'package_composition_change', mas é "
                f"'{alert.alert_type}'."
            )

        # Verificar que os atributos reportados correspondem
        alerted_attrs = {a.attribute_name for a in alerts}
        assert alerted_attrs == set(attrs_to_change), (
            f"Atributos alertados deveriam ser "
            f"{set(attrs_to_change)}, mas são {alerted_attrs}."
        )

    @settings(max_examples=100)
    @given(comp_data=identical_composition_pair_strategy())
    def test_property_11_previous_none_baseline_zero_alertas(
        self, comp_data: dict
    ) -> None:
        """Property 11: Primeiro registro (previous=None) → 0 alertas.

        Quando previous é None (primeiro registro = baseline),
        _compare_compositions SHALL retornar lista vazia sem gerar
        alertas, tratando o registro como baseline.

        **Validates: Requirements 7.1, 7.3**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()
        plan_name = comp_data["plan_name"]
        attrs = {
            k: v for k, v in comp_data.items() if k != "plan_name"
        }

        current_pkg = _make_package_mock(plan_name, attrs)

        # previous=None (baseline)
        alerts = detector._compare_compositions(
            current=[current_pkg],
            previous=None,
        )

        assert len(alerts) == 0, (
            f"Primeiro registro (previous=None) deveria gerar "
            f"0 alertas (baseline), mas gerou {len(alerts)}. "
            f"Plan: '{plan_name}'"
        )

    @settings(max_examples=100)
    @given(comp_data=identical_composition_pair_strategy())
    def test_property_11_previous_lista_vazia_baseline_zero_alertas(
        self, comp_data: dict
    ) -> None:
        """Property 11: Primeiro registro (previous=[]) → 0 alertas.

        Quando previous é lista vazia, _compare_compositions SHALL
        retornar lista vazia (equivalente a baseline).

        **Validates: Requirements 7.1, 7.3**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()
        plan_name = comp_data["plan_name"]
        attrs = {
            k: v for k, v in comp_data.items() if k != "plan_name"
        }

        current_pkg = _make_package_mock(plan_name, attrs)

        # previous=[] (baseline)
        alerts = detector._compare_compositions(
            current=[current_pkg],
            previous=[],
        )

        assert len(alerts) == 0, (
            f"Previous lista vazia deveria gerar 0 alertas "
            f"(baseline), mas gerou {len(alerts)}. "
            f"Plan: '{plan_name}'"
        )

    @settings(max_examples=100)
    @given(pkg_data=new_package_strategy())
    def test_property_11_pacote_novo_alerta_por_atributo_nao_null(
        self, pkg_data: dict
    ) -> None:
        """Property 11: Pacote novo → alerta para cada atributo não-null.

        Quando um pacote está presente em current mas não em
        previous (plan_name novo), _compare_compositions SHALL
        gerar alertas para cada atributo não-null do pacote novo,
        com previous_value=None e current_value preenchido.

        **Validates: Requirements 7.1, 7.3**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()
        plan_name = pkg_data["plan_name"]
        attrs = {
            k: v for k, v in pkg_data.items() if k != "plan_name"
        }

        current_pkg = _make_package_mock(plan_name, attrs)

        # Previous contém um pacote com plan_name diferente
        other_pkg = _make_package_mock(
            "Outro Plano Existente", {"default_price": 100.0}
        )

        alerts = detector._compare_compositions(
            current=[current_pkg],
            previous=[other_pkg],
        )

        # Contar atributos não-null no pacote novo
        non_null_attrs = [
            attr for attr in _COMPOSITION_ATTRS_FOR_COMPARISON
            if attrs.get(attr) is not None
        ]
        expected_count = len(non_null_attrs)

        assert len(alerts) == expected_count, (
            f"Pacote novo deveria gerar {expected_count} alertas "
            f"(1 por atributo não-null), mas gerou {len(alerts)}. "
            f"Atributos não-null: {non_null_attrs}. "
            f"Plan: '{plan_name}'"
        )

        # Verificar tipo de todos os alertas
        for alert in alerts:
            assert alert.alert_type == "package_composition_change", (
                f"Tipo do alerta deveria ser "
                f"'package_composition_change', mas é "
                f"'{alert.alert_type}'."
            )

        # Verificar previous_value=None para pacote novo
        for alert in alerts:
            assert alert.previous_value is None, (
                f"Pacote novo deveria ter previous_value=None, "
                f"mas tem '{alert.previous_value}' para "
                f"atributo '{alert.attribute_name}'."
            )

        # Verificar current_value não é None
        for alert in alerts:
            assert alert.current_value is not None, (
                f"Pacote novo deveria ter current_value preenchido, "
                f"mas é None para atributo "
                f"'{alert.attribute_name}'."
            )

    @settings(max_examples=100)
    @given(
        comp_data=identical_composition_pair_strategy(),
        attrs_to_change=changed_attributes_strategy(),
    )
    def test_property_11_alertas_contem_valores_anterior_atual(
        self, comp_data: dict, attrs_to_change: frozenset
    ) -> None:
        """Property 11: Alertas contêm previous_value e current_value.

        Para qualquer atributo alterado, o alerta gerado SHALL
        conter previous_value com o valor anterior (como string)
        e current_value com o valor atual (como string).

        **Validates: Requirements 7.1, 7.3**
        """
        from price_watchdog.comparator.change_detector import (
            ChangeDetector,
        )

        detector = ChangeDetector()
        plan_name = comp_data["plan_name"]

        # Construir atributos do registro anterior (base)
        previous_attrs = {
            k: v for k, v in comp_data.items() if k != "plan_name"
        }

        # Construir atributos do registro atual (com mudanças)
        current_attrs = dict(previous_attrs)
        for attr_name in attrs_to_change:
            original_val = previous_attrs.get(attr_name)
            if attr_name in ("default_price", "promotional_price"):
                if original_val is None:
                    current_attrs[attr_name] = 299.99
                else:
                    current_attrs[attr_name] = original_val + 100.0
            elif attr_name == "promotional_period_months":
                if original_val is None:
                    current_attrs[attr_name] = 12
                else:
                    current_attrs[attr_name] = (
                        (original_val % 36) + 1
                    )
            elif attr_name in (
                "linear_channels",
                "simultaneous_screens",
                "fiber_speed_mbps",
                "mobile_speed_mbps",
            ):
                if original_val is None:
                    current_attrs[attr_name] = 200
                else:
                    current_attrs[attr_name] = original_val + 50
            elif attr_name in ("has_fiber", "has_mobile_internet"):
                if original_val is None:
                    current_attrs[attr_name] = True
                else:
                    current_attrs[attr_name] = not original_val
            elif attr_name in (
                "bundled_streaming_1",
                "bundled_streaming_2",
                "bundled_streaming_3",
            ):
                if original_val is None:
                    current_attrs[attr_name] = "Paramount+"
                elif original_val == "Paramount+":
                    current_attrs[attr_name] = "Disney+"
                else:
                    current_attrs[attr_name] = "Paramount+"

        previous_pkg = _make_package_mock(plan_name, previous_attrs)
        current_pkg = _make_package_mock(plan_name, current_attrs)

        alerts = detector._compare_compositions(
            current=[current_pkg],
            previous=[previous_pkg],
        )

        # Verificar que cada alerta contém valores corretos
        for alert in alerts:
            attr = alert.attribute_name
            prev_val = previous_attrs.get(attr)
            curr_val = current_attrs.get(attr)

            # previous_value e current_value são strings ou None
            expected_prev = (
                str(prev_val) if prev_val is not None else None
            )
            expected_curr = (
                str(curr_val) if curr_val is not None else None
            )

            assert alert.previous_value == expected_prev, (
                f"Atributo '{attr}': previous_value deveria ser "
                f"'{expected_prev}', mas é '{alert.previous_value}'."
            )
            assert alert.current_value == expected_curr, (
                f"Atributo '{attr}': current_value deveria ser "
                f"'{expected_curr}', mas é '{alert.current_value}'."
            )


# --- Estratégias (generators) para Property 13 ---


def _competitor_name_strategy() -> st.SearchStrategy[str]:
    """Gera nomes de concorrentes válidos."""
    return st.text(
        alphabet=st.characters(
            whitelist_categories=("L", "N", "Z"),
            blacklist_characters="\x00",
        ),
        min_size=2,
        max_size=30,
    ).filter(lambda s: s.strip() != "")


def _package_mock_for_report() -> st.SearchStrategy[dict]:
    """Gera dados de pacote para geração de relatório Excel.

    Retorna um dicionário com todos os atributos de um
    PackageComposition para uso em mocks de relatório.
    Usa texto seguro para Excel (sem caracteres de controle).
    """
    return st.fixed_dictionaries({
        "plan_name": _excel_safe_text(min_size=1, max_size=40),
        "default_price": st.one_of(
            st.none(),
            st.floats(
                min_value=10.0,
                max_value=5000.0,
                allow_nan=False,
                allow_infinity=False,
            ),
        ),
        "promotional_price": st.one_of(
            st.none(),
            st.floats(
                min_value=10.0,
                max_value=4999.0,
                allow_nan=False,
                allow_infinity=False,
            ),
        ),
        "promotional_period_months": st.one_of(
            st.none(),
            st.integers(min_value=1, max_value=36),
        ),
        "linear_channels": st.one_of(
            st.none(),
            st.integers(min_value=0, max_value=500),
        ),
        "simultaneous_screens": st.one_of(
            st.none(),
            st.integers(min_value=1, max_value=10),
        ),
        "has_fiber": st.one_of(st.none(), st.booleans()),
        "fiber_speed_mbps": st.one_of(
            st.none(),
            st.integers(min_value=100, max_value=2000),
        ),
        "has_mobile_internet": st.one_of(st.none(), st.booleans()),
        "mobile_speed_mbps": st.one_of(
            st.none(),
            st.integers(min_value=50, max_value=1000),
        ),
        "bundled_streaming_1": st.one_of(
            st.none(),
            st.sampled_from(["Netflix", "Disney+", "Paramount+"]),
        ),
        "bundled_streaming_2": st.one_of(
            st.none(),
            st.sampled_from(["Globoplay", "HBO Max", "Star+"]),
        ),
        "bundled_streaming_3": st.one_of(
            st.none(),
            st.sampled_from(["Apple TV+", "Amazon Prime Video"]),
        ),
    })


def _excel_safe_text(
    min_size: int = 1, max_size: int = 50
) -> st.SearchStrategy[str]:
    """Gera texto seguro para uso em células Excel (sem control chars).

    openpyxl rejeita caracteres de controle (U+0000-U+001F exceto
    tab/newline, e outros ilegais em XML). Usamos apenas categorias
    seguras: letras, números, pontuação e espaços.
    """
    return st.text(
        alphabet=st.characters(
            whitelist_categories=("L", "N", "P", "Z"),
            blacklist_characters="\x00\x01\x02\x03\x04\x05\x06"
            "\x07\x08\x0b\x0c\x0e\x0f\x10\x11\x12\x13\x14\x15"
            "\x16\x17\x18\x19\x1a\x1b\x1c\x1d\x1e\x1f",
        ),
        min_size=min_size,
        max_size=max_size,
    ).filter(lambda s: s.strip() != "")


def _intelligence_record_for_report() -> st.SearchStrategy[dict]:
    """Gera dados para um CompetitorIntelligenceRecord com status success.

    Cada record contém 1 a 5 pacotes e dados de comunicação
    comercial para geração de relatório.
    Usa texto seguro para Excel (sem caracteres de controle).
    """
    return st.fixed_dictionaries({
        "competitor_name": _competitor_name_strategy(),
        "packages": st.lists(
            _package_mock_for_report(),
            min_size=1,
            max_size=5,
        ),
        "commercial_keywords": st.one_of(
            st.none(),
            st.lists(
                _excel_safe_text(min_size=1, max_size=40),
                min_size=3,
                max_size=10,
            ),
        ),
        "home_banner_description": st.one_of(
            st.none(),
            _excel_safe_text(min_size=5, max_size=200),
        ),
        "commercial_positioning_summary": st.one_of(
            st.none(),
            _excel_safe_text(min_size=5, max_size=300),
        ),
    })


def _build_report_record_mock(data: dict) -> MagicMock:
    """Constrói mock de CompetitorIntelligenceRecord para relatório.

    Args:
        data: Dicionário com dados do record gerados pela strategy.

    Returns:
        MagicMock com atributos simulando um record de sucesso.
    """
    from price_watchdog.models.intelligence_entities import (
        CompetitorIntelligenceRecord,
        PackageComposition,
    )

    record = MagicMock(spec=CompetitorIntelligenceRecord)
    record.extraction_status = "success"
    record.competitor = MagicMock()
    record.competitor.name = data["competitor_name"]
    record.commercial_keywords = data["commercial_keywords"]
    record.home_banner_description = data["home_banner_description"]
    record.commercial_positioning_summary = data[
        "commercial_positioning_summary"
    ]

    # Construir mocks de pacotes
    pkg_mocks = []
    for pkg_data in data["packages"]:
        pkg = MagicMock(spec=PackageComposition)
        pkg.plan_name = pkg_data["plan_name"]
        pkg.default_price = pkg_data["default_price"]
        pkg.promotional_price = pkg_data["promotional_price"]
        pkg.promotional_period_months = pkg_data[
            "promotional_period_months"
        ]
        pkg.linear_channels = pkg_data["linear_channels"]
        pkg.simultaneous_screens = pkg_data["simultaneous_screens"]
        pkg.has_fiber = pkg_data["has_fiber"]
        pkg.fiber_speed_mbps = pkg_data["fiber_speed_mbps"]
        pkg.has_mobile_internet = pkg_data["has_mobile_internet"]
        pkg.mobile_speed_mbps = pkg_data["mobile_speed_mbps"]
        pkg.bundled_streaming_1 = pkg_data["bundled_streaming_1"]
        pkg.bundled_streaming_2 = pkg_data["bundled_streaming_2"]
        pkg.bundled_streaming_3 = pkg_data["bundled_streaming_3"]
        pkg_mocks.append(pkg)

    record.packages = pkg_mocks
    return record


def intelligence_records_for_report_strategy() -> (
    st.SearchStrategy[list[dict]]
):
    """Gera lista não-vazia de records com status success para relatório."""
    return st.lists(
        _intelligence_record_for_report(),
        min_size=1,
        max_size=5,
    )


# --- Property 13 Tests ---


@pytest.mark.property
class TestExcelReportIntelligenceTabsProperty:
    """Property 13: Relatório Excel contém abas de inteligência com estrutura correta.

    Para qualquer conjunto não-vazio de CompetitorIntelligenceRecords
    com status "success", o ExcelReportGenerator SHALL produzir um
    arquivo Excel contendo: uma aba "Composição de Pacotes" com uma
    linha por pacote e colunas [Concorrente, Nome do Pacote, Preço
    Default, Preço Promocional, Duração Promo, Canais Lineares, Telas
    Simultâneas, Fibra, Velocidade Fibra, Internet Móvel, Velocidade
    Móvel, Streaming 1, Streaming 2, Streaming 3], e uma aba
    "Comunicação Comercial" com colunas [Concorrente, Palavras-chave,
    Descrição Banner, Resumo Posicionamento].

    Feature: competitor-intelligence-expansion, Property 13

    **Validates: Requirements 6.1, 6.2, 6.3**
    """

    @settings(max_examples=100, deadline=None)
    @given(records_data=intelligence_records_for_report_strategy())
    def test_property_13_composicao_tab_existe_com_14_colunas(
        self, records_data: list[dict]
    ) -> None:
        """Property 13: Aba "Composição de Pacotes" existe com 14 colunas.

        Para qualquer conjunto de records com sucesso, a aba
        "Composição de Pacotes" SHALL existir no Workbook e
        ter exatamente 14 colunas no cabeçalho.

        **Validates: Requirements 6.1, 6.2**
        """
        from openpyxl import Workbook

        from price_watchdog.reports.excel_report import (
            ExcelReportGenerator,
        )

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _build_report_record_mock(d) for d in records_data
        ]

        generator._generate_composition_tab(wb, records)

        # Verificar que a aba existe
        assert "Composição de Pacotes" in wb.sheetnames, (
            f"Aba 'Composição de Pacotes' deveria existir no "
            f"Workbook, mas abas presentes são: {wb.sheetnames}"
        )

        ws = wb["Composição de Pacotes"]

        # Verificar que tem 14 colunas no cabeçalho (linha 1)
        header_values = [
            ws.cell(row=1, column=col).value
            for col in range(1, 15)
        ]
        non_none_headers = [h for h in header_values if h is not None]
        assert len(non_none_headers) == 14, (
            f"Cabeçalho deveria ter 14 colunas, mas tem "
            f"{len(non_none_headers)}. Headers: {header_values}"
        )

        # Verificar que a coluna 15 está vazia (apenas 14 colunas)
        assert ws.cell(row=1, column=15).value is None, (
            "Coluna 15 deveria estar vazia (apenas 14 colunas)."
        )

    @settings(max_examples=100, deadline=None)
    @given(records_data=intelligence_records_for_report_strategy())
    def test_property_13_comunicacao_tab_existe_com_4_colunas(
        self, records_data: list[dict]
    ) -> None:
        """Property 13: Aba "Comunicação Comercial" existe com 4 colunas.

        Para qualquer conjunto de records com sucesso, a aba
        "Comunicação Comercial" SHALL existir e ter exatamente
        4 colunas no cabeçalho.

        **Validates: Requirements 6.3**
        """
        from openpyxl import Workbook

        from price_watchdog.reports.excel_report import (
            ExcelReportGenerator,
        )

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _build_report_record_mock(d) for d in records_data
        ]

        generator._generate_communication_tab(wb, records)

        # Verificar que a aba existe
        assert "Comunicação Comercial" in wb.sheetnames, (
            f"Aba 'Comunicação Comercial' deveria existir no "
            f"Workbook, mas abas presentes são: {wb.sheetnames}"
        )

        ws = wb["Comunicação Comercial"]

        # Verificar que tem 4 colunas no cabeçalho (linha 1)
        header_values = [
            ws.cell(row=1, column=col).value
            for col in range(1, 5)
        ]
        non_none_headers = [h for h in header_values if h is not None]
        assert len(non_none_headers) == 4, (
            f"Cabeçalho deveria ter 4 colunas, mas tem "
            f"{len(non_none_headers)}. Headers: {header_values}"
        )

        # Verificar que a coluna 5 está vazia (apenas 4 colunas)
        assert ws.cell(row=1, column=5).value is None, (
            "Coluna 5 deveria estar vazia (apenas 4 colunas)."
        )

    @settings(max_examples=100, deadline=None)
    @given(records_data=intelligence_records_for_report_strategy())
    def test_property_13_composicao_rows_igual_total_pacotes(
        self, records_data: list[dict]
    ) -> None:
        """Property 13: Linhas de dados = total de pacotes nos records.

        Para qualquer conjunto de records com sucesso, o número
        de linhas de dados na aba "Composição de Pacotes" SHALL
        ser igual ao total de pacotes em todos os records.

        **Validates: Requirements 6.1, 6.2**
        """
        from openpyxl import Workbook

        from price_watchdog.reports.excel_report import (
            ExcelReportGenerator,
        )

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _build_report_record_mock(d) for d in records_data
        ]

        generator._generate_composition_tab(wb, records)

        ws = wb["Composição de Pacotes"]

        # Contar total de pacotes esperados
        total_packages = sum(
            len(d["packages"]) for d in records_data
        )

        # Contar linhas de dados (a partir da linha 2, após cabeçalho)
        data_rows = 0
        row = 2
        while ws.cell(row=row, column=1).value is not None or \
              ws.cell(row=row, column=2).value is not None:
            data_rows += 1
            row += 1

        assert data_rows == total_packages, (
            f"Aba 'Composição de Pacotes' deveria ter "
            f"{total_packages} linhas de dados (1 por pacote), "
            f"mas tem {data_rows}. "
            f"Records: {len(records_data)}"
        )

    @settings(max_examples=100, deadline=None)
    @given(records_data=intelligence_records_for_report_strategy())
    def test_property_13_comunicacao_rows_igual_num_records(
        self, records_data: list[dict]
    ) -> None:
        """Property 13: Linhas de dados na aba comunicação = num records.

        Para qualquer conjunto de records com sucesso, o número
        de linhas de dados na aba "Comunicação Comercial" SHALL
        ser igual ao número de records com sucesso.

        **Validates: Requirements 6.3**
        """
        from openpyxl import Workbook

        from price_watchdog.reports.excel_report import (
            ExcelReportGenerator,
        )

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _build_report_record_mock(d) for d in records_data
        ]

        generator._generate_communication_tab(wb, records)

        ws = wb["Comunicação Comercial"]

        # Todos os records gerados têm status "success"
        expected_rows = len(records_data)

        # Contar linhas de dados (a partir da linha 2, após cabeçalho)
        data_rows = 0
        row = 2
        while ws.cell(row=row, column=1).value is not None:
            data_rows += 1
            row += 1

        assert data_rows == expected_rows, (
            f"Aba 'Comunicação Comercial' deveria ter "
            f"{expected_rows} linhas de dados (1 por record "
            f"com sucesso), mas tem {data_rows}."
        )
