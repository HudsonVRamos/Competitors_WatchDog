"""Property-based tests para completude do relatório Excel.

Feature: price-watchdog, Property 15: Relatório Excel contém todos os
records do ciclo.

Validates: Requirements 10.1
"""

from datetime import datetime, timezone
from io import BytesIO

import pytest
from hypothesis import given, settings
from hypothesis.strategies import (
    composite,
    floats,
    integers,
    lists,
    text,
)
from openpyxl import load_workbook

from price_watchdog.reports.excel_report import (
    ExcelReportGenerator,
    ReportRow,
)


# Colunas obrigatórias do relatório
REQUIRED_COLUMNS = [
    "Concorrente",
    "Produto",
    "Nosso Preço",
    "Preço Deles",
    "Diferença (R$)",
    "Diferença (%)",
    "Status",
]

# Estratégias para dados realistas
positive_prices = floats(
    min_value=0.01,
    max_value=100_000.0,
    allow_nan=False,
    allow_infinity=False,
)

percentage_values = floats(
    min_value=-100.0,
    max_value=500.0,
    allow_nan=False,
    allow_infinity=False,
)

# Texto não-vazio para nomes (sem \x00 que pode quebrar Excel)
non_empty_text = text(
    min_size=1,
    max_size=50,
    alphabet="abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 áéíóúãõçÁÉÍÓÚÃÕÇ",
)


@composite
def report_row_strategy(draw):
    """Gera um ReportRow com dados realistas."""
    our_price = draw(positive_prices)
    extracted_price = draw(positive_prices)
    difference = extracted_price - our_price
    if our_price > 0:
        difference_pct = (extracted_price - our_price) / our_price * 100
    else:
        difference_pct = 0.0

    return ReportRow(
        competitor_name=draw(non_empty_text),
        product_name=draw(non_empty_text),
        our_price=our_price,
        extracted_price=extracted_price,
        price_difference=difference,
        price_difference_pct=difference_pct,
    )


class _FakeCycle:
    """Substituto leve de PriceCycle para evitar instrumentação SQLAlchemy."""

    def __init__(self):
        self.started_at = datetime(2024, 1, 15, 10, 0, 0, tzinfo=timezone.utc)


def _make_cycle():
    """Cria um ciclo mínimo para testes (sem SQLAlchemy)."""
    return _FakeCycle()


def _load_workbook_from_bytes(excel_bytes: bytes):
    """Carrega workbook openpyxl a partir de bytes."""
    return load_workbook(BytesIO(excel_bytes))


@pytest.mark.property
class TestReportCompletenessProperties:
    """Testes de propriedade para completude do relatório Excel.

    **Validates: Requirements 10.1**
    """

    @given(
        rows=lists(
            report_row_strategy(),
            min_size=1,
            max_size=50,
        )
    )
    @settings(max_examples=100, deadline=None)
    def test_report_contains_exact_number_of_data_rows(
        self, rows: list[ReportRow]
    ) -> None:
        """Property 15: Número de linhas de dados == número de ReportRows.

        **Validates: Requirements 10.1**

        Para qualquer conjunto de N ReportRows, o relatório Excel
        gerado deve conter exatamente N linhas de dados (a partir
        da linha 5, após o cabeçalho na linha 4).
        """
        generator = ExcelReportGenerator()
        cycle = _make_cycle()

        excel_bytes = generator.generate_from_rows(rows, cycle)
        wb = _load_workbook_from_bytes(excel_bytes)
        ws = wb.active

        # Dados começam na linha 5, cabeçalho na linha 4
        data_rows = 0
        for row_idx in range(5, ws.max_row + 1):
            # Verifica se a linha tem conteúdo (coluna 1 não vazia)
            if ws.cell(row=row_idx, column=1).value is not None:
                data_rows += 1

        assert data_rows == len(rows), (
            f"Esperado {len(rows)} linhas de dados, "
            f"encontrado {data_rows}"
        )

    @given(
        rows=lists(
            report_row_strategy(),
            min_size=1,
            max_size=20,
        )
    )
    @settings(max_examples=100, deadline=None)
    def test_report_contains_all_required_columns(
        self, rows: list[ReportRow]
    ) -> None:
        """Property 15: Todas as 7 colunas obrigatórias presentes.

        **Validates: Requirements 10.1**

        O relatório deve conter exatamente as colunas: Concorrente,
        Produto, Nosso Preço, Preço Deles, Diferença (R$),
        Diferença (%), Status.
        """
        generator = ExcelReportGenerator()
        cycle = _make_cycle()

        excel_bytes = generator.generate_from_rows(rows, cycle)
        wb = _load_workbook_from_bytes(excel_bytes)
        ws = wb.active

        # Cabeçalhos estão na linha 4
        header_values = []
        for col_idx in range(1, len(REQUIRED_COLUMNS) + 1):
            cell_value = ws.cell(row=4, column=col_idx).value
            header_values.append(cell_value)

        for expected_col in REQUIRED_COLUMNS:
            assert expected_col in header_values, (
                f"Coluna obrigatória '{expected_col}' não encontrada "
                f"no cabeçalho. Colunas presentes: {header_values}"
            )

    @given(
        rows=lists(
            report_row_strategy(),
            min_size=1,
            max_size=30,
        )
    )
    @settings(max_examples=100, deadline=None)
    def test_each_row_contains_correct_data(
        self, rows: list[ReportRow]
    ) -> None:
        """Property 15: Cada linha contém dados corretos do ReportRow.

        **Validates: Requirements 10.1**

        Cada linha de dados no Excel deve corresponder ao ReportRow
        de entrada na mesma posição, com os valores corretos em
        cada coluna.
        """
        generator = ExcelReportGenerator()
        cycle = _make_cycle()

        excel_bytes = generator.generate_from_rows(rows, cycle)
        wb = _load_workbook_from_bytes(excel_bytes)
        ws = wb.active

        for idx, row_data in enumerate(rows):
            excel_row = 5 + idx  # dados começam na linha 5

            # Coluna 1: Concorrente
            assert ws.cell(row=excel_row, column=1).value == row_data.competitor_name, (
                f"Linha {idx}: competitor_name incorreto"
            )

            # Coluna 2: Produto
            assert ws.cell(row=excel_row, column=2).value == row_data.product_name, (
                f"Linha {idx}: product_name incorreto"
            )

            # Coluna 3: Nosso Preço
            assert ws.cell(row=excel_row, column=3).value == pytest.approx(
                row_data.our_price, rel=1e-9
            ), (
                f"Linha {idx}: our_price incorreto"
            )

            # Coluna 4: Preço Deles
            assert ws.cell(row=excel_row, column=4).value == pytest.approx(
                row_data.extracted_price, rel=1e-9
            ), (
                f"Linha {idx}: extracted_price incorreto"
            )

            # Coluna 5: Diferença (R$)
            assert ws.cell(row=excel_row, column=5).value == pytest.approx(
                row_data.price_difference, rel=1e-9
            ), (
                f"Linha {idx}: price_difference incorreto"
            )

            # Coluna 6: Diferença (%)
            assert ws.cell(row=excel_row, column=6).value == pytest.approx(
                row_data.price_difference_pct, rel=1e-9
            ), (
                f"Linha {idx}: price_difference_pct incorreto"
            )

            # Coluna 7: Status (deve ser um dos 3 valores válidos)
            status_value = ws.cell(row=excel_row, column=7).value
            assert status_value in (
                "Competitivo",
                "Atenção",
                "Não Competitivo",
            ), (
                f"Linha {idx}: status '{status_value}' inválido"
            )

    @given(
        rows=lists(
            report_row_strategy(),
            min_size=0,
            max_size=0,
        )
    )
    @settings(max_examples=10, deadline=None)
    def test_empty_report_has_no_data_rows(
        self, rows: list[ReportRow]
    ) -> None:
        """Property 15: Relatório com 0 records gera 0 linhas de dados.

        **Validates: Requirements 10.1**

        Caso especial: quando não há records, o relatório deve conter
        apenas o cabeçalho sem linhas de dados.
        """
        generator = ExcelReportGenerator()
        cycle = _make_cycle()

        excel_bytes = generator.generate_from_rows(rows, cycle)
        wb = _load_workbook_from_bytes(excel_bytes)
        ws = wb.active

        # Não deve haver dados a partir da linha 5
        for row_idx in range(5, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            assert cell_value is None, (
                f"Encontrado dado na linha {row_idx} em relatório vazio"
            )
