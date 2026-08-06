"""Testes unitários para geração de aba 'Comunicação Comercial' no relatório Excel.

Valida que o ExcelReportGenerator gera corretamente a aba de comunicação
comercial com as colunas especificadas e uma linha por concorrente com
extração bem-sucedida.
"""

from __future__ import annotations

from io import BytesIO
from unittest.mock import MagicMock

from openpyxl import load_workbook

from price_watchdog.reports.excel_report import ExcelReportGenerator


def _make_intelligence_record(
    competitor_name: str,
    extraction_status: str = "success",
    keywords: list[str] | None = None,
    banner_description: str | None = None,
    positioning_summary: str | None = None,
):
    """Cria um mock de CompetitorIntelligenceRecord para testes."""
    record = MagicMock()
    record.extraction_status = extraction_status
    record.competitor = MagicMock()
    record.competitor.name = competitor_name
    record.commercial_keywords = keywords
    record.home_banner_description = banner_description
    record.commercial_positioning_summary = positioning_summary
    record.packages = []
    return record


class TestCommunicationTab:
    """Testes para o método _generate_communication_tab."""

    def test_generates_correct_headers(self):
        """Aba deve conter cabeçalhos corretos na primeira linha."""
        from openpyxl import Workbook

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _make_intelligence_record(
                "ConcorrenteA",
                keywords=["oferta", "fibra", "streaming"],
                banner_description="Banner test",
                positioning_summary="Resumo test",
            )
        ]

        generator._generate_communication_tab(wb, records)

        ws = wb["Comunicação Comercial"]
        assert ws.cell(row=1, column=1).value == "Concorrente"
        assert ws.cell(row=1, column=2).value == "Palavras-chave"
        assert ws.cell(row=1, column=3).value == "Descrição Banner"
        assert ws.cell(row=1, column=4).value == "Resumo Posicionamento"

    def test_one_row_per_successful_competitor(self):
        """Deve gerar uma linha por concorrente com status success."""
        from openpyxl import Workbook

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _make_intelligence_record(
                "ConcorrenteA",
                keywords=["oferta", "fibra", "streaming"],
                banner_description="Banner A",
                positioning_summary="Resumo A",
            ),
            _make_intelligence_record(
                "ConcorrenteB",
                extraction_status="failed",
                keywords=None,
                banner_description=None,
                positioning_summary=None,
            ),
            _make_intelligence_record(
                "ConcorrenteC",
                keywords=["desconto", "promo", "grátis"],
                banner_description="Banner C",
                positioning_summary="Resumo C",
            ),
        ]

        generator._generate_communication_tab(wb, records)

        ws = wb["Comunicação Comercial"]
        # Apenas 2 registros com success (A e C)
        assert ws.cell(row=2, column=1).value == "ConcorrenteA"
        assert ws.cell(row=3, column=1).value == "ConcorrenteC"
        # Linha 4 não deve ter dados
        assert ws.cell(row=4, column=1).value is None

    def test_keywords_joined_by_comma(self):
        """Palavras-chave devem ser separadas por vírgula."""
        from openpyxl import Workbook

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _make_intelligence_record(
                "ConcorrenteA",
                keywords=["oferta", "fibra óptica", "streaming grátis"],
                banner_description="Banner",
                positioning_summary="Resumo",
            )
        ]

        generator._generate_communication_tab(wb, records)

        ws = wb["Comunicação Comercial"]
        assert (
            ws.cell(row=2, column=2).value
            == "oferta, fibra óptica, streaming grátis"
        )

    def test_empty_keywords_renders_empty_string(self):
        """Keywords None ou vazio devem resultar em string vazia."""
        from openpyxl import Workbook

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _make_intelligence_record(
                "ConcorrenteA",
                keywords=None,
                banner_description="Banner",
                positioning_summary="Resumo",
            )
        ]

        generator._generate_communication_tab(wb, records)

        ws = wb["Comunicação Comercial"]
        assert ws.cell(row=2, column=2).value == ""

    def test_null_fields_render_as_empty_string(self):
        """Campos null devem resultar em strings vazias."""
        from openpyxl import Workbook

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _make_intelligence_record(
                "ConcorrenteA",
                keywords=["a", "b", "c"],
                banner_description=None,
                positioning_summary=None,
            )
        ]

        generator._generate_communication_tab(wb, records)

        ws = wb["Comunicação Comercial"]
        assert ws.cell(row=2, column=3).value == ""
        assert ws.cell(row=2, column=4).value == ""

    def test_filters_only_success_status(self):
        """Apenas registros com extraction_status='success' aparecem."""
        from openpyxl import Workbook

        generator = ExcelReportGenerator()
        wb = Workbook()
        records = [
            _make_intelligence_record(
                "Failed", extraction_status="failed"
            ),
            _make_intelligence_record(
                "NoPackages", extraction_status="no_packages_found"
            ),
            _make_intelligence_record(
                "Success",
                extraction_status="success",
                keywords=["test", "kw", "three"],
                banner_description="Banner",
                positioning_summary="Resumo",
            ),
        ]

        generator._generate_communication_tab(wb, records)

        ws = wb["Comunicação Comercial"]
        assert ws.cell(row=2, column=1).value == "Success"
        assert ws.cell(row=3, column=1).value is None

    def test_no_records_generates_only_headers(self):
        """Sem registros, a aba deve ter apenas os cabeçalhos."""
        from openpyxl import Workbook

        generator = ExcelReportGenerator()
        wb = Workbook()

        generator._generate_communication_tab(wb, [])

        ws = wb["Comunicação Comercial"]
        assert ws.cell(row=1, column=1).value == "Concorrente"
        assert ws.cell(row=2, column=1).value is None
